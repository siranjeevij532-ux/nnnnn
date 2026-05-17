import re
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils import timezone
from django.db.models import Sum, ExpressionWrapper, DecimalField, F as F_db, Avg as models_Avg

# ── GF(256) lookup tables for QR code generation ─────────────────────────────
GF256_EXP = [0] * 512
GF256_LOG  = [0] * 256
_gfx = 1
for _gfi in range(255):
    GF256_EXP[_gfi] = _gfx
    GF256_LOG[_gfx] = _gfi
    _gfx <<= 1
    if _gfx & 0x100: _gfx ^= 0x11d
for _gfi in range(255, 512): GF256_EXP[_gfi] = GF256_EXP[_gfi - 255]
# ─────────────────────────────────────────────────────────────────────────────
from .models import Table, Category, MenuItem, Order, OrderItem, ShopSettings, Discount, CustomerProfile, Combo, ComboItem, PosDraft
import json
import openpyxl
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP


# ═══════════════════════════════════════════════════════════════════
# CUSTOMER LOGIN & HISTORY
# ═══════════════════════════════════════════════════════════════════

def get_water_bottle(order_type=None):
    """Return Water Bottle MenuItem for cart page quick-add."""
    qs = MenuItem.objects.filter(name__icontains='water bottle')
    if order_type == 'takeaway':
        qs = qs.filter(is_available_takeaway=True)
    elif order_type == 'dine_in':
        qs = qs.filter(is_available_dine_in=True)
    bottle = qs.order_by('id').first()
    if bottle:
        return bottle

    # Fallback to any water-related item if no exact water bottle is found.
    qs = MenuItem.objects.filter(name__icontains='water')
    if order_type == 'takeaway':
        qs = qs.filter(is_available_takeaway=True)
    elif order_type == 'dine_in':
        qs = qs.filter(is_available_dine_in=True)
    bottle = qs.order_by('id').first()
    if bottle:
        return bottle

    # Fallback to any active beverage item.
    qs = MenuItem.objects.filter(item_type='beverage')
    if order_type == 'takeaway':
        qs = qs.filter(is_available_takeaway=True)
    elif order_type == 'dine_in':
        qs = qs.filter(is_available_dine_in=True)
    bottle = qs.order_by('id').first()
    if bottle:
        return bottle

    # Create a default Water Bottle item if nothing exists yet.
    category = Category.objects.filter(is_active=True).first()
    if not category:
        category = Category.objects.create(name='Beverages', icon='🥤', description='Drinks and refreshments', order=0, is_active=True)
    bottle = MenuItem.objects.create(
        category=category,
        name='Water Bottle',
        description='Pure drinking water',
        price=10,
        item_type='beverage',
        is_available_dine_in=True,
        is_available_takeaway=True,
        is_featured=False,
        preparation_time=1,
        order=0,
        parcel_charge=0,
    )
    return bottle


def get_shop():
    return ShopSettings.objects.first()

def guest_login(request):
    """Guest login — no registration needed. Sets a guest session and redirects."""
    import random as _rand
    next_url = request.GET.get('next', '/')
    # Assign a unique guest phone so orders can be tracked within the session
    guest_phone = f"GUEST{_rand.randint(100000, 999999)}"
    request.session['customer_phone'] = guest_phone
    request.session['customer_name']  = 'Guest'
    request.session['customer_id']    = None
    request.session['is_guest']       = True
    return redirect(next_url)


def customer_login(request):
    """
    Combined Register / Login page.
    Register: name + phone → creates CustomerProfile, sets session.
    Login: phone only → looks up existing profile, sets session.
    """
    next_url = request.GET.get('next', '/')
    shop = get_shop()

    # Already logged in → go to home
    if request.session.get('customer_phone'):
        return redirect('/customer/home/')

    context = {'next': next_url, 'shop': shop}

    if request.method == 'POST':
        action = request.POST.get('action', 'register')
        next_url = request.POST.get('next', '/')

        if action == 'register':
            name  = request.POST.get('name', '').strip()
            phone = request.POST.get('phone', '').strip()
            context['reg_name']  = name
            context['reg_phone'] = phone

            if not name:
                context['reg_error'] = 'Please enter your name.'
            elif not phone or len(phone) < 8:
                context['reg_error'] = 'Please enter a valid phone number.'
            elif CustomerProfile.objects.filter(phone=phone).exists():
                context['reg_error'] = 'This phone number is already registered. Please use Login.'
            else:
                profile = CustomerProfile.objects.create(name=name, phone=phone)
                request.session['customer_phone'] = phone
                request.session['customer_name']  = name
                request.session['customer_id']    = profile.id
                return redirect(next_url)

        elif action == 'login':
            phone = request.POST.get('phone', '').strip()
            context['login_phone'] = phone

            if not phone or len(phone) < 8:
                context['login_error'] = 'Please enter a valid phone number.'
            else:
                try:
                    profile = CustomerProfile.objects.get(phone=phone)
                    profile.visit_count += 1
                    profile.save()
                    request.session['customer_phone'] = phone
                    request.session['customer_name']  = profile.name
                    request.session['customer_id']    = profile.id
                    return redirect(next_url)
                except CustomerProfile.DoesNotExist:
                    context['login_error'] = 'Phone number not found. Please register first.'

    return render(request, 'restaurant/customer_login.html', context)


def customer_logout(request):
    request.session.flush()
    return redirect('/')


def customer_history(request):
    phone = request.session.get('customer_phone')
    if not phone:
        return redirect("/customer/login/?next=/customer/history/")
    shop = get_shop()
    name = request.session.get('customer_name', '')

    all_orders = Order.objects.filter(
        customer_phone=phone,
        parent_order__isnull=True,
    ).prefetch_related('items__menu_item','reorders__items__menu_item').order_by('-created_at')

    def enrich(order):
        linked = [order] + list(order.reorders.order_by('created_at'))
        combined_total = sum(o.total_amount for o in linked)
        all_items = []
        for o in linked:
            for item in o.items.select_related('menu_item').all():
                item.order = o
                all_items.append(item)
        return {
            'order': order,
            'combined_total': combined_total,
            'all_items': all_items,
            'all_orders': linked,      # for live card to show per-order grouping
            'reorder_count': len(linked)-1,
        }

    # Active = any non-completed/non-cancelled root order
    ACTIVE_STATUSES = ['pending','accepted','preparing','ready']
    active_entry = None
    past_entries = []
    for o in all_orders:
        e = enrich(o)
        if o.status in ACTIVE_STATUSES and active_entry is None:
            active_entry = e
        else:
            past_entries.append(e)

    all_entries = [enrich(o) for o in all_orders]

    # Stats
    completed_count = sum(1 for o in all_orders if o.status == 'completed')
    from django.db.models import Sum
    total_spent = Order.objects.filter(
        customer_phone=phone,
        status='completed',
    ).aggregate(t=Sum('total_amount'))['t'] or 0

    # Last table - prefer session value set when menu was loaded (for refresh support)
    last_order = all_orders.first()
    sess_table = request.session.get('last_table_id')
    sess_type  = request.session.get('last_order_type', '')
    if sess_table:
        last_table_id = sess_table
        is_takeaway   = False
    elif sess_type == 'takeaway':
        last_table_id = None
        is_takeaway   = True
    else:
        last_table_id = last_order.table.id if last_order and last_order.table else None
        is_takeaway   = last_order.order_type == 'takeaway' if last_order else False

    prog_steps = [
        ('pending',   '📋', 'Received'),
        ('accepted',  '✅', 'Accepted'),
        ('preparing', '👨‍🍳', 'Cooking'),
        ('ready',     '🔔', 'Ready'),
        ('completed', '🎉', 'Done'),
    ]

    return render(request, 'restaurant/customer_home.html', {
        'orders': all_entries,
        'active_order': active_entry,
        'past_orders': past_entries,
        'customer_name': name,
        'customer_phone': phone,
        'completed_count': completed_count,
        'total_spent': total_spent,
        'last_table_id': last_table_id,
        'is_takeaway': is_takeaway,
        'prog_steps': prog_steps,
        'shop': shop,
    })


def table_selection(request):
    """Root URL — redirect customers to login, show nothing useful."""
    # If already logged in, show a simple welcome with history link
    if request.session.get('customer_phone'):
        name = request.session.get('customer_name','')
        return redirect('/customer/history/')
    return redirect('/customer/login/?next=/customer/history/')


def _get_active_order_for_menu(phone):
    """Return active order entry for a customer phone, or None."""
    if not phone:
        return None
    ACTIVE_STATUSES = ['pending', 'accepted', 'preparing', 'ready']
    order = Order.objects.filter(
        customer_phone=phone,
        parent_order__isnull=True,
        status__in=ACTIVE_STATUSES,
    ).prefetch_related('items__menu_item', 'reorders__items__menu_item').order_by('-created_at').first()
    if not order:
        return None
    linked = [order] + list(order.reorders.order_by('created_at'))
    combined_subtotal = sum(o.subtotal for o in linked)
    combined_parcel   = sum(o.parcel_charge for o in linked)
    combined_discount = sum(o.discount_amount for o in linked)
    combined_total    = combined_subtotal + combined_parcel - combined_discount
    return {'order': order, 'combined_total': combined_total, 'combined_parcel': combined_parcel, 'all_orders': linked}


def _get_order_history_for_menu(phone, limit=5):
    """Return past orders for a customer (completed/cancelled), newest first."""
    if not phone:
        return []
    past = Order.objects.filter(
        customer_phone=phone,
        parent_order__isnull=True,
        status__in=['completed', 'cancelled'],
    ).prefetch_related('items__menu_item', 'reorders__items__menu_item').order_by('-created_at')[:limit]
    result = []
    for o in past:
        linked = [o] + list(o.reorders.order_by('created_at'))
        combined_total = sum(x.subtotal + x.parcel_charge - x.discount_amount for x in linked)
        all_items = []
        for x in linked:
            for item in x.items.select_related('menu_item').all():
                all_items.append(item)
        result.append({'order': o, 'combined_total': combined_total, 'all_items': all_items, 'reorder_count': len(linked) - 1})
    return result


@ensure_csrf_cookie
def menu_view(request, table_id):
    # Require customer login (skip for reorders — session already set)
    reorder_from = request.GET.get('reorder_from')
    if not request.session.get('customer_phone') and not reorder_from:
        return redirect(f"/customer/login/?next=/table/{table_id}/menu/")

    table = get_object_or_404(Table, id=table_id, is_active=True)
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts = Discount.objects.filter(is_active=True)
    shop = get_shop()
    reorder_order = None
    reorder_customer_name = ''
    reorder_customer_phone = ''
    cart = '{}'
    cart_count = 0
    if reorder_from:
        try:
            # Walk up to root order so we always link to the original
            reorder_order = Order.objects.get(id=reorder_from)
            while reorder_order.parent_order:
                reorder_order = reorder_order.parent_order
            reorder_customer_name = reorder_order.customer_name
            reorder_customer_phone = reorder_order.customer_phone
            prefill = {
                str(item.menu_item.id): {
                    'name': item.menu_item.name,
                    'price': float(item.unit_price),
                    'qty': item.quantity,
                }
                for item in reorder_order.items.select_related('menu_item').all()
            }
            if prefill:
                cart = json.dumps(prefill)
                cart_count = sum(v['qty'] for v in prefill.values())
        except Order.DoesNotExist:
            reorder_order = None

    phone = request.session.get('customer_phone', '')
    active_order = _get_active_order_for_menu(phone)
    past_orders = _get_order_history_for_menu(phone)
    combos = Combo.objects.filter(is_active=True).prefetch_related('combo_items__menu_item')

    # Store last menu URL in session so page refresh works
    request.session['last_table_id'] = table_id
    request.session['last_order_type'] = 'dine_in'

    wb = get_water_bottle(order_type='dine_in')
    return render(request, 'restaurant/menu.html', {
        'table': table, 'categories': categories,
        'discounts': discounts, 'shop': shop,
        'reorder_from': str(reorder_order.id) if reorder_order else None,
        'reorder_order': reorder_order,
        'cart': cart,
        'cart_count': cart_count,
        'is_takeaway': False,
        'active_order': active_order,
        'past_orders': past_orders,
        'combos': combos,
        'water_bottle_id': wb.id if wb else None,
        'water_bottle_name': wb.name if wb else 'Water Bottle',
        'water_bottle_price': wb.price if wb else 10,
    })


@ensure_csrf_cookie
def takeaway_menu(request):
    """Takeaway order — no table needed"""
    if not request.session.get('customer_phone'):
        return redirect("/customer/login/?next=/takeaway/")
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts = Discount.objects.filter(is_active=True)
    shop = get_shop()
    phone = request.session.get('customer_phone', '')
    active_order = _get_active_order_for_menu(phone)
    past_orders = _get_order_history_for_menu(phone)
    combos = Combo.objects.filter(is_active=True).prefetch_related('combo_items__menu_item')
    request.session['last_table_id'] = None
    request.session['last_order_type'] = 'takeaway'
    wb = get_water_bottle(order_type='takeaway')
    # Build parcel charges map for frontend: {item_id: parcel_charge}
    default_pc = shop.default_parcel_charge if shop else 0
    parcel_map = {}
    for cat in categories:
        for item in cat.items.all():
            pc = float(item.parcel_charge) if item.parcel_charge else float(default_pc)
            parcel_map[item.id] = pc  # include all items, 0 means no charge

    import json as _json
    return render(request, 'restaurant/menu.html', {
        'table': None, 'categories': categories,
        'discounts': discounts, 'shop': shop,
        'cart': '{}', 'cart_count': 0,
        'is_takeaway': True,
        'active_order': active_order,
        'past_orders': past_orders,
        'combos': combos,
        'water_bottle_id': wb.id if wb else None,
        'water_bottle_name': wb.name if wb else 'Water Bottle',
        'water_bottle_price': wb.price if wb else 10,
        'default_parcel_charge': float(default_pc),
        'parcel_map_json': _json.dumps(parcel_map),
    })


def place_order(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
        customer_name  = (data.get('customer_name') or request.session.get('customer_name', '')).strip()
        customer_phone = (data.get('customer_phone') or request.session.get('customer_phone', '')).strip()
        special_instructions = data.get('special_instructions', '')
        items = data.get('items', [])
        discount_percent = Decimal(str(data.get('discount_percent', 0)))
        order_type = data.get('order_type', 'dine_in')  # 'dine_in' or 'takeaway'
        table_id = data.get('table_id')
        parent_order_id = data.get('parent_order_id')
        reorder_from_id = data.get('reorder_from')  # parent order id for reorders

        if not customer_name:
            return JsonResponse({'success': False, 'error': 'Customer name is required'})
        if not customer_phone:
            return JsonResponse({'success': False, 'error': 'Phone number is required'})
        if not items:
            return JsonResponse({'success': False, 'error': 'No items in order'})

        table = None
        if order_type == 'dine_in' and table_id:
            table = Table.objects.filter(id=table_id, is_active=True).first()
            if not table:
                return JsonResponse({'success': False, 'error': 'Table not found'})

        special_note = special_instructions
        parent_order_obj = None
        if parent_order_id:
            try:
                parent_order_obj = Order.objects.get(id=parent_order_id)
                special_note = f'Reorder (linked to #{parent_order_obj.order_number})' + (f' — {special_instructions}' if special_instructions else '')
            except Order.DoesNotExist:
                pass

        import random
        order = Order(
            table=table,
            customer_name=customer_name,
            customer_phone=customer_phone,
            special_instructions=special_note,
            discount_percent=discount_percent,
            order_type=order_type,
            order_number=f"BC{timezone.now().strftime('%Y%m%d')}{random.randint(1000,9999)}",
            subtotal=Decimal('0'),
            discount_amount=Decimal('0'),
            total_amount=Decimal('0'),
            parent_order=parent_order_obj,
        )
        order.save()

        subtotal = Decimal('0')

        # Batch-fetch all needed menu items and combos in 2 queries (not N queries)
        regular_ids = [item_data['id'] for item_data in items
                       if not item_data.get('combo_id') and item_data.get('id')]
        combo_ids   = [item_data['combo_id'] for item_data in items if item_data.get('combo_id')]

        # Filter items by availability based on order_type
        if order_type == 'takeaway':
            menu_items_map = {m.id: m for m in MenuItem.objects.filter(id__in=regular_ids, is_available_takeaway=True)} if regular_ids else {}
            fallback_item  = MenuItem.objects.filter(is_available_takeaway=True).first() if combo_ids else None
        else:  # dine_in
            menu_items_map = {m.id: m for m in MenuItem.objects.filter(id__in=regular_ids, is_available_dine_in=True)} if regular_ids else {}
            fallback_item  = MenuItem.objects.filter(is_available_dine_in=True).first() if combo_ids else None
        
        combos_map     = {c.id: c for c in Combo.objects.filter(id__in=combo_ids, is_active=True).prefetch_related('combo_items__menu_item')} if combo_ids else {}

        order_items_to_create = []
        for item_data in items:
            combo_id = item_data.get('combo_id')
            if combo_id:
                combo = combos_map.get(int(combo_id))
                if not combo:
                    continue
                qty = max(1, int(item_data.get('qty', 1)))
                combo_price = combo.price * qty
                combo_items = list(combo.combo_items.all())
                rep_item = combo_items[0].menu_item if combo_items else fallback_item
                if rep_item:
                    order_items_to_create.append(OrderItem(
                        order=order, menu_item=rep_item, quantity=qty,
                        unit_price=combo.price, notes=f'🎁 Combo: {combo.name}'
                    ))
                subtotal += combo_price
                continue
            item_id = item_data.get('id')
            if not item_id:
                continue
            menu_item = menu_items_map.get(int(item_id))
            if not menu_item:
                continue
            qty = max(1, int(item_data.get('qty', 1)))
            order_items_to_create.append(OrderItem(
                order=order, menu_item=menu_item, quantity=qty, unit_price=menu_item.price
            ))
            subtotal += menu_item.price * qty

        # Single bulk insert instead of N inserts
        OrderItem.objects.bulk_create(order_items_to_create)

        # Add parcel charges for takeaway orders
        parcel_total = Decimal('0')
        if order_type == 'takeaway':
            shop_obj = get_shop()
            default_pc = shop_obj.default_parcel_charge if shop_obj else Decimal('0')
            for oi in order_items_to_create:
                item_parcel = oi.menu_item.parcel_charge if oi.menu_item.parcel_charge else default_pc
                parcel_total += item_parcel * oi.quantity

        # subtotal = items only (no parcel), parcel stored separately
        order.subtotal = subtotal
        order.parcel_charge = parcel_total
        if subtotal > 0 and discount_percent > 0:
            order.discount_amount = (subtotal * discount_percent / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            order.discount_amount = Decimal('0.00')
        order.total_amount = order.subtotal + order.parcel_charge - order.discount_amount
        Order.objects.filter(pk=order.pk).update(
            subtotal=order.subtotal,
            parcel_charge=order.parcel_charge,
            discount_amount=order.discount_amount,
            total_amount=order.total_amount,
        )
        if table:
            Table.objects.filter(pk=table.pk).update(status='occupied')

        return JsonResponse({'success': True, 'order_id': order.id, 'order_number': order.order_number, 'history_url': '/customer/history/'})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid request data'}, status=400)
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@ensure_csrf_cookie
def reorder_menu(request, order_id):
    """Show menu so customer can pick new items — these will be added to existing order bill"""
    order = get_object_or_404(Order, id=order_id)
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts = Discount.objects.filter(is_active=True)
    shop = get_shop()
    import json as _json
    prefill = {str(item.menu_item.id): {'name': item.menu_item.name, 'price': float(item.unit_price), 'qty': item.quantity}
               for item in order.items.all()}
    return render(request, 'restaurant/menu.html', {
        'table': order.table,
        'categories': categories,
        'discounts': discounts,
        'shop': shop,
        'cart': _json.dumps(prefill),
        'cart_count': sum(v['qty'] for v in prefill.values()),
        'reorder_for': order,
        'is_reorder': True,
    })


def add_items_to_order(request, order_id):
    """Add new items from reorder menu into the existing order bill"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        order = get_object_or_404(Order, id=order_id)
        data = json.loads(request.body)
        items = data.get('items', [])
        order_type = data.get('order_type', order.order_type)  # Use order's type if not specified
        if not items:
            return JsonResponse({'success': False, 'error': 'No items provided'})
        for item_data in items:
            combo_id = item_data.get('combo_id')
            if combo_id:
                try:
                    combo = Combo.objects.prefetch_related('combo_items__menu_item').get(id=combo_id, is_active=True)
                except Combo.DoesNotExist:
                    continue
                qty = max(1, int(item_data.get('qty', 1)))
                combo_price = combo.price * qty
                combo_items = list(combo.combo_items.select_related('menu_item').all())
                if combo_items:
                    raw_total = sum(ci.menu_item.price * ci.quantity for ci in combo_items)
                    for ci in combo_items:
                        if raw_total > 0:
                            proportion = (ci.menu_item.price * ci.quantity) / raw_total
                            unit = (combo_price * proportion / ci.quantity).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        else:
                            unit = (combo_price / len(combo_items) / ci.quantity).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        OrderItem.objects.create(
                            order=order, menu_item=ci.menu_item,
                            quantity=ci.quantity * qty, unit_price=unit,
                            notes=f'[Combo: {combo.name}]'
                        )
                continue
            menu_item = MenuItem.objects.filter(id=item_data['id']).first()
            if not menu_item:
                continue
            # Check availability based on order type
            if order_type == 'takeaway' and not menu_item.is_available_takeaway:
                continue
            if order_type == 'dine_in' and not menu_item.is_available_dine_in:
                continue
            qty = max(1, int(item_data.get('qty', 1)))
            existing = order.items.filter(menu_item=menu_item).first()
            if existing:
                existing.quantity += qty
                existing.save()
            else:
                OrderItem.objects.create(order=order, menu_item=menu_item,
                                         quantity=qty, unit_price=menu_item.price)
        order.calculate_totals()
        Order.objects.filter(pk=order.pk).update(
            subtotal=order.subtotal,
            discount_amount=order.discount_amount,
            total_amount=order.total_amount,
        )
        return JsonResponse({'success': True, 'order_id': order.id, 'order_number': order.order_number, 'history_url': '/customer/history/'})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def reorder(request, order_id):
    """Redirect customer to menu with original items pre-loaded for reorder"""
    orig = get_object_or_404(Order, id=order_id)
    if orig.table:
        return redirect(f"/table/{orig.table.id}/menu/?reorder_from={orig.id}")
    else:
        return redirect(f"/takeaway/?reorder_from={orig.id}")


def order_status(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    shop = get_shop()

    # Find root order and gather all orders in this session (first + reorders)
    root_order = order
    while root_order.parent_order:
        root_order = root_order.parent_order
    all_orders = [root_order] + list(root_order.reorders.order_by('created_at'))

    # Build combined items list and total across all orders
    combined_items = []
    combined_subtotal = Decimal('0')
    for o in all_orders:
        for item in o.items.select_related('menu_item').all():
            combined_items.append({
                'name': item.menu_item.name,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price,
                'is_reorder': o.parent_order is not None,
            })
            combined_subtotal += item.total_price
    discount_percent = root_order.discount_percent
    discount_amount = (combined_subtotal * discount_percent / 100).quantize(Decimal('0.01')) if discount_percent else Decimal('0')
    combined_total = combined_subtotal - discount_amount

    # Bill is ready only when every order in session is completed
    all_completed = all(o.status == 'completed' for o in all_orders)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': order.status,           # this order's status (for progress bar)
            'root_status': root_order.status, # root order status
            'all_completed': all_completed,   # True only when EVERY order is done
            'payment_status': order.payment_status,
            'status_display': order.get_status_display(),
            'combined_total': float(combined_total),
            'has_reorders': len(all_orders) > 1,
        })
    return render(request, 'restaurant/order_status.html', {
        'order': order,
        'root_order': root_order,
        'all_orders': all_orders,
        'combined_items': combined_items,
        'combined_subtotal': combined_subtotal,
        'discount_percent': discount_percent,
        'discount_amount': discount_amount,
        'combined_total': combined_total,
        'has_reorders': len(all_orders) > 1,
        'all_completed': all_completed,
        'shop': shop,
    })


def bill_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    shop = get_shop()
    # Find root order - bill is only available when root order is completed
    root_order = order
    while root_order.parent_order:
        root_order = root_order.parent_order
    all_orders_check = [root_order] + list(root_order.reorders.order_by('created_at'))
    all_done = all(o.status == 'completed' for o in all_orders_check)
    if not all_done:
        return render(request, 'restaurant/waiting_for_bill.html', {'order': root_order, 'shop': shop})
    all_orders = [root_order] + list(root_order.reorders.order_by('created_at'))
    combined_items = []
    for o in all_orders:
        for item in o.items.select_related('menu_item').all():
            combined_items.append({
                'order_number': o.order_number,
                'name': item.menu_item.name,
                'notes': item.notes,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price,
                'is_reorder': o.parent_order is not None,
            })
    # Recalculate totals properly (handles old orders where total_amount may not include parcel)
    combined_subtotal = sum(o.subtotal for o in all_orders)
    # Recover parcel from special_instructions for old orders
    def get_order_parcel(o):
        if o.parcel_charge:
            return o.parcel_charge
        # Old orders stored parcel in special_instructions
        m = re.search(r'\[Parcel Charge: ₹([\d.]+)\]', o.special_instructions or '')
        from decimal import Decimal as D
        return D(m.group(1)) if m else D('0')
    combined_parcel   = sum(get_order_parcel(o) for o in all_orders)
    discount_percent  = root_order.discount_percent
    discount_amount   = sum(o.discount_amount for o in all_orders)
    combined_total    = combined_subtotal + combined_parcel - discount_amount
    return render(request, 'restaurant/bill.html', {
        'order': order,
        'root_order': root_order,
        'all_orders': all_orders,
        'combined_items': combined_items,
        'combined_subtotal': combined_subtotal,
        'combined_parcel': combined_parcel,
        'discount_percent': discount_percent,
        'discount_amount': discount_amount,
        'combined_total': combined_total,
        'has_reorders': len(all_orders) > 1,
        'shop': shop,
    })


def pay_online(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    shop = get_shop()
    order.payment_method = 'online'
    order.payment_status = 'online_pending'
    order.save()
    return render(request, 'restaurant/pay_online.html', {'order': order, 'shop': shop})


def confirm_payment(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if request.method == 'POST':
        order.payment_status = 'paid_online'
        order.payment_method = 'online'
        order.completed_at = timezone.now()
        order.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})


def download_bill(request, order_id):
    """Keep for backward compat but redirect to bill view"""
    return redirect('bill', order_id=order_id)




def print_bill(request, order_id):
    """Thermal printer-friendly bill page for staff"""
    order = get_object_or_404(Order, id=order_id)
    shop = get_shop()
    root_order = order
    while root_order.parent_order:
        root_order = root_order.parent_order
    all_orders = [root_order] + list(root_order.reorders.order_by('created_at'))
    combined_items = []
    combined_subtotal = Decimal('0')
    for o in all_orders:
        for item in o.items.select_related('menu_item').all():
            combined_items.append({
                'name': item.menu_item.name,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price,
            })
            combined_subtotal += item.total_price
    discount_percent = root_order.discount_percent
    discount_amount = (combined_subtotal * discount_percent / 100).quantize(Decimal('0.01')) if discount_percent else Decimal('0')
    combined_total = combined_subtotal - discount_amount
    return render(request, 'restaurant/print_bill.html', {
        'order': order,
        'root_order': root_order,
        'all_orders': all_orders,
        'combined_items': combined_items,
        'combined_subtotal': combined_subtotal,
        'discount_percent': discount_percent,
        'discount_amount': discount_amount,
        'combined_total': combined_total,
        'shop': shop,
    })



def reorder_notification_api(request):
    """Returns new reorder (child) orders since last_id for staff sound alert"""
    try:
        last_id = int(request.GET.get('last_id', 0))
        orders = Order.objects.filter(
            id__gt=last_id,
            parent_order__isnull=False,
            status='pending'
        ).select_related('table', 'parent_order')
        result = []
        for o in orders:
            result.append({
                'id': o.id,
                'order_number': o.order_number,
                'customer_name': o.customer_name,
                'table': o.table.name if o.table else 'Takeaway',
                'total': float(o.total_amount),
                'parent_order_number': o.parent_order.order_number if o.parent_order else '',
                'is_reorder': True,
            })
        return JsonResponse({'orders': result})
    except Exception as e:
        return JsonResponse({'orders': [], 'error': str(e)})

# ─── STAFF PORTAL ─────────────────────────────────────────────────

def staff_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user and user.is_staff:
            login(request, user)
            return redirect('staff_portal')
        return render(request, 'restaurant/staff_login.html', {'error': 'Invalid credentials'})
    return render(request, 'restaurant/staff_login.html')


def staff_logout(request):
    logout(request)
    return redirect('staff_login')



# ═══════════════════════════════════════════════════════════════════
# LIVE ORDER PAGE
# ═══════════════════════════════════════════════════════════════════

def live_order(request):
    """Dedicated live order page — reads order_id from query param or session."""
    if not request.session.get('customer_phone'):
        return redirect('/customer/login/')
    return render(request, 'restaurant/live_order.html', {})


def live_order_data(request):
    """JSON API for the live order page — returns full order state."""
    order_id = request.GET.get('order_id')
    if not order_id:
        return JsonResponse({'error': 'no order id'}, status=400)
    try:
        root = Order.objects.get(id=order_id)
        # Walk to root
        while root.parent_order:
            root = root.parent_order
    except Order.DoesNotExist:
        return JsonResponse({'error': 'not found'}, status=404)

    all_orders = [root] + list(root.reorders.order_by('created_at'))
    all_completed = all(o.status == 'completed' for o in all_orders)

    orders_data = []
    for o in all_orders:
        items_data = []
        for item in o.items.select_related('menu_item').all():
            items_data.append({
                'name': item.notes if item.notes else item.menu_item.name,
                'qty':  item.quantity,
                'total': float(item.total_price),
            })
        orders_data.append({'items': items_data, 'total': float(o.total_amount)})

    # Recalculate totals correctly (subtotal + parcel - discount)
    combined_subtotal = sum(o.subtotal for o in all_orders)
    combined_parcel   = sum(o.parcel_charge for o in all_orders)
    combined_discount = sum(o.discount_amount for o in all_orders)
    combined_total    = combined_subtotal + combined_parcel - combined_discount

    return JsonResponse({
        'order_id':        root.id,
        'order_number':    root.order_number,
        'status':          root.status,
        'all_completed':   all_completed,
        'customer_name':   root.customer_name,
        'customer_phone':  root.customer_phone,
        'combined_total':  float(combined_total),
        'combined_parcel': float(combined_parcel),
        'orders':          orders_data,
    })


def combo_detail_api(request, combo_id):
    """JSON API — returns combo details with items for customer modal."""
    try:
        combo = Combo.objects.prefetch_related('combo_items__menu_item').get(id=combo_id, is_active=True)
    except Combo.DoesNotExist:
        return JsonResponse({'error': 'not found'}, status=404)
    items_data = []
    for ci in combo.combo_items.all():
        mi = ci.menu_item
        items_data.append({
            'name': mi.name,
            'qty': ci.quantity,
            'price': float(mi.price),
            'image': mi.image.url if mi.image else None,
            'type': mi.item_type,
        })
    return JsonResponse({
        'id': combo.id,
        'name': combo.name,
        'description': combo.description,
        'price': float(combo.price),
        'icon': combo.icon or '🎁',
        'image': combo.image.url if combo.image else None,
        'items': items_data,
    })



# ── POS DRAFT VIEWS ────────────────────────────────────────────

@login_required(login_url='/staff/login/')
def pos_save_draft(request):
    """Save a POS draft order to the database (includes slot, parcel, order_type)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        import json as _json
        from decimal import Decimal as D
        data = _json.loads(request.body)
        name       = (data.get('name') or 'Walk-in').strip()
        phone      = (data.get('phone') or '').strip()
        slot       = (data.get('slot') or '').strip()
        table      = (data.get('table') or slot or 'Takeaway').strip()
        items      = data.get('items', [])
        sub        = D(str(data.get('subtotal', 0)))
        parcel     = D(str(data.get('parcel', 0)))
        disc       = D(str(data.get('discount_pct', 0)))
        total      = D(str(data.get('total', 0)))
        note       = (data.get('note') or '').strip()
        order_type = (data.get('order_type') or 'dine_in').strip()
        draft_id   = data.get('draft_id')
        table_name = slot if slot else table

        if draft_id:
            try:
                draft = PosDraft.objects.get(id=draft_id)
                draft.customer_name  = name
                draft.customer_phone = phone
                draft.table_name     = table_name
                draft.items_json     = _json.dumps(items)
                draft.subtotal       = sub
                draft.discount_pct   = disc
                draft.total_amount   = total
                draft.note           = note
                # Store extra fields in note if model doesn't have them
                draft.save()
            except PosDraft.DoesNotExist:
                draft_id = None

        if not draft_id:
            import random
            num = f"D{timezone.now().strftime('%Y%m%d')}{random.randint(100,999)}"
            draft = PosDraft.objects.create(
                draft_number=num, customer_name=name, customer_phone=phone,
                table_name=table_name, items_json=_json.dumps(items),
                subtotal=sub, discount_pct=disc, total_amount=total, note=note
            )

        return JsonResponse({
            'success': True,
            'draft_id': draft.id,
            'draft_number': draft.draft_number,
            'slot': slot,
            'order_type': order_type,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def pos_get_drafts(request):
    """Return all active POS drafts (includes slot and order_type derived from table_name)."""
    import json as _json
    drafts = PosDraft.objects.filter(is_deleted=False).order_by('-created_at')[:50]
    result = []
    for d in drafts:
        table_name = d.table_name or 'Takeaway'
        is_tw = table_name == 'Takeaway' or table_name.startswith('Takeaway ')
        try:
            raw_items = _json.loads(d.items_json)
        except Exception:
            raw_items = []
        result.append({
            'id': d.id,
            'draft_number': d.draft_number,
            'customer_name': d.customer_name,
            'customer_phone': d.customer_phone,
            'table_name': table_name,
            'slot': table_name,
            'order_type': 'takeaway' if is_tw else 'dine_in',
            'total_amount': float(d.total_amount),
            'subtotal': float(d.subtotal),
            'parcel': 0,  # parcel not stored separately in current schema
            'items': raw_items,
            'discount_pct': float(d.discount_pct),
            'note': d.note,
            'created_at': d.created_at.strftime('%d %b %Y %H:%M'),
        })
    return JsonResponse({'drafts': result})


@login_required(login_url='/staff/login/')
def pos_delete_draft(request, draft_id):
    """Soft-delete a POS draft."""
    try:
        d = PosDraft.objects.get(id=draft_id)
        d.is_deleted = True
        d.save()
        return JsonResponse({'success': True})
    except PosDraft.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'}, status=404)


def pos_save_draft_no_auth(request):
    """Save a POS draft order to the database without authentication requirement."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        import json as _json
        from decimal import Decimal as D
        data = _json.loads(request.body)
        name       = (data.get('name') or 'Walk-in').strip()
        phone      = (data.get('phone') or '').strip()
        slot       = (data.get('slot') or '').strip()
        table      = (data.get('table') or slot or 'Takeaway').strip()
        items      = data.get('items', [])
        sub        = D(str(data.get('subtotal', 0)))
        parcel     = D(str(data.get('parcel', 0)))
        disc       = D(str(data.get('discount_pct', 0)))
        total      = D(str(data.get('total', 0)))
        note       = (data.get('note') or '').strip()
        order_type = (data.get('order_type') or 'dine_in').strip()
        draft_id   = data.get('draft_id')
        table_name = slot if slot else table

        if draft_id:
            try:
                draft = PosDraft.objects.get(id=draft_id)
                draft.customer_name  = name
                draft.customer_phone = phone
                draft.table_name     = table_name
                draft.items_json     = _json.dumps(items)
                draft.subtotal       = sub
                draft.discount_pct   = disc
                draft.total_amount   = total
                draft.note           = note
                draft.save()
            except PosDraft.DoesNotExist:
                draft_id = None

        if not draft_id:
            import random
            num = f"D{timezone.now().strftime('%Y%m%d')}{random.randint(100,999)}"
            draft = PosDraft.objects.create(
                draft_number=num, customer_name=name, customer_phone=phone,
                table_name=table_name, items_json=_json.dumps(items),
                subtotal=sub, discount_pct=disc, total_amount=total, note=note
            )

        return JsonResponse({'success': True, 'draft_id': draft.id, 'draft_number': draft.draft_number})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def pos_cart_live_push(request):
    """Receives the full POS cart (all slots) and upserts PosDraft records
    so the staff portal live sidebar can display them in real time.
    Called from POS terminal every time items change (addItem / removeItem).
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        import json as _json
        data = _json.loads(request.body)
        slots = data.get('slots', {})   # { "Table 1": {items, disc, isTW, parcel, name}, ... }

        # Mark sentinel so we can tell live-cart drafts from real drafts
        LIVE_MARKER = '__LIVE__'

        for slot_name, cart in slots.items():
            items = cart.get('items', [])
            # Build subtotal
            sub = sum(i.get('price', 0) * i.get('qty', 1) for i in items)
            disc_pct = float(cart.get('disc', 0) or 0)
            parcel = float(cart.get('parcel', 0) or 0)
            dis = round(sub * disc_pct / 100, 2)
            total = round(sub + parcel - dis, 2)

            # Upsert: find existing live-cart draft for this slot, or create
            draft = PosDraft.objects.filter(
                table_name=slot_name,
                note__startswith=LIVE_MARKER,
                is_deleted=False
            ).first()

            items_json = _json.dumps(items)
            if draft:
                draft.items_json   = items_json
                draft.subtotal     = sub
                draft.discount_pct = disc_pct
                draft.total_amount = total
                draft.customer_name = cart.get('name') or 'Walk-in'
                draft.save()
            else:
                import random
                num = f"LC{timezone.now().strftime('%Y%m%d%H%M%S')}{random.randint(10,99)}"
                PosDraft.objects.create(
                    draft_number=num,
                    customer_name=cart.get('name') or 'Walk-in',
                    customer_phone=cart.get('phone') or '',
                    table_name=slot_name,
                    items_json=items_json,
                    subtotal=sub,
                    discount_pct=disc_pct,
                    total_amount=total,
                    note=f"{LIVE_MARKER}{slot_name}",
                )

        # Delete live-cart drafts for slots that are now empty / removed
        active_slots = set(slots.keys())
        PosDraft.objects.filter(
            note__startswith=LIVE_MARKER,
            is_deleted=False
        ).exclude(table_name__in=active_slots).update(is_deleted=True)

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def pos_save_order(request):
    """Save a POS order directly into the staff portal Order table.
    No customer session required. Staff login required.
    Status = completed + paid so it goes straight to Excel, no staff portal popup.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        import random as _random
        data = json.loads(request.body)

        name         = (data.get('customer_name') or 'Walk-in').strip() or 'Walk-in'
        phone        = (data.get('customer_phone') or '0000000000').strip() or '0000000000'
        items_data   = data.get('items', [])
        disc_pct     = Decimal(str(data.get('discount_percent', 0) or 0))
        order_type   = (data.get('order_type') or 'dine_in').strip()
        slot         = (data.get('slot') or '').strip()
        parcel_total = Decimal(str(data.get('parcel_total', 0) or 0))

        if not items_data:
            return JsonResponse({'success': False, 'error': 'No items in order'})

        # Pre-fetch all regular menu items in one query
        regular_ids = [int(d['id']) for d in items_data if d.get('id')]
        menu_items_map = {m.id: m for m in MenuItem.objects.filter(id__in=regular_ids)} if regular_ids else {}

        # Pre-fetch combo items
        combo_ids = [int(d['combo_id']) for d in items_data if d.get('combo_id')]
        combos_map = {c.id: c for c in Combo.objects.filter(id__in=combo_ids, is_active=True)} if combo_ids else {}

        # Create the order
        order = Order(
            customer_name=name,
            customer_phone=phone,
            special_instructions=f'[POS][Slot:{slot}]' if slot else '[POS]',
            discount_percent=disc_pct,
            order_type=order_type,
            order_number=f"POS{timezone.now().strftime('%Y%m%d')}{_random.randint(1000, 9999)}",
            subtotal=Decimal('0'),
            discount_amount=Decimal('0'),
            total_amount=Decimal('0'),
            status='completed',
            payment_status='paid',
        )
        order.save()

        subtotal = Decimal('0')
        to_create = []

        for d in items_data:
            qty        = max(1, int(d.get('qty', 1)))
            sent_price = Decimal(str(d.get('price', 0) or 0))

            if d.get('combo_id'):
                # Combo item — use combo's representative first item or sent price
                combo = combos_map.get(int(d['combo_id']))
                if not combo:
                    # No DB record — create a placeholder using first available item
                    fallback = MenuItem.objects.filter(is_available=True).first()
                    if not fallback:
                        continue
                    unit_price = sent_price if sent_price > 0 else Decimal('0')
                    to_create.append(OrderItem(
                        order=order,
                        menu_item=fallback,
                        quantity=qty,
                        unit_price=unit_price,
                        notes=f'Combo: {d.get("name", "?")}',
                    ))
                    subtotal += unit_price * qty
                    continue
                # Use first combo item's menu_item as the representative row
                first_ci = combo.combo_items.select_related('menu_item').first()
                rep_item = first_ci.menu_item if first_ci else MenuItem.objects.filter(is_available=True).first()
                if not rep_item:
                    continue
                unit_price = sent_price if sent_price > 0 else combo.price
                to_create.append(OrderItem(
                    order=order,
                    menu_item=rep_item,
                    quantity=qty,
                    unit_price=unit_price,
                    notes=f'Combo: {combo.name}',
                ))
                subtotal += unit_price * qty

            else:
                # Regular item
                mid = d.get('id')
                if not mid:
                    continue
                menu_item = menu_items_map.get(int(mid))
                if not menu_item:
                    continue
                # Use sent price (reflects any parcel/discount already shown on POS)
                unit_price = sent_price if sent_price > 0 else menu_item.price
                to_create.append(OrderItem(
                    order=order,
                    menu_item=menu_item,
                    quantity=qty,
                    unit_price=unit_price,
                ))
                subtotal += unit_price * qty

        OrderItem.objects.bulk_create(to_create)

        disc_amount  = (subtotal * disc_pct / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if disc_pct else Decimal('0')
        total        = subtotal + parcel_total - disc_amount

        Order.objects.filter(pk=order.pk).update(
            subtotal=subtotal,
            discount_amount=disc_amount,
            total_amount=total,
            parcel_charge=parcel_total,
        )

        # Clean up the live-cart draft for this slot so sidebar removes it
        if slot:
            PosDraft.objects.filter(
                table_name=slot,
                note__startswith='__LIVE__',
                is_deleted=False
            ).update(is_deleted=True)

        return JsonResponse({
            'success': True,
            'order_id': order.id,
            'order_number': order.order_number,
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def staff_portal(request):
    from django.db.models import Count, Avg
    tables = Table.objects.filter(is_active=True).order_by('number')
    pending_orders  = Order.objects.filter(status='pending').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    accepted_orders = Order.objects.filter(status='accepted').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    cooking_orders  = Order.objects.filter(status='preparing').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    ready_orders    = Order.objects.filter(status='ready').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    completed_orders = Order.objects.filter(status='completed').select_related('table').prefetch_related('items__menu_item').order_by('-created_at')[:30]

    def _fix_totals(qs):
        """Annotate each order with real total calculated from items (fixes stale DB values)."""
        result = list(qs)
        for o in result:
            items = list(o.items.all())
            sub = sum(i.unit_price * i.quantity for i in items)
            disc = (sub * o.discount_percent / 100).quantize(Decimal('0.01')) if o.discount_percent else Decimal('0')
            o.total_amount = sub - disc
        return result

    pending_orders  = _fix_totals(pending_orders)
    accepted_orders = _fix_totals(accepted_orders)
    cooking_orders  = _fix_totals(cooking_orders)
    ready_orders    = _fix_totals(ready_orders)
    completed_orders = _fix_totals(completed_orders)

    today = timezone.now().date()
    week_start  = today - timezone.timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    def _revenue_from_items(order_qs):
        """Calculate true revenue from OrderItem rows to avoid stale total_amount=0 in DB."""
        result = OrderItem.objects.filter(order__in=order_qs).aggregate(
            rev=Sum(ExpressionWrapper(F_db('unit_price') * F_db('quantity'), output_field=DecimalField()))
        )['rev'] or Decimal('0')
        # Subtract discounts
        total = Decimal('0')
        for o in order_qs.only('subtotal','discount_percent','discount_amount','total_amount','id'):
            sub = sum(i.unit_price * i.quantity for i in o.items.all())
            disc = (sub * o.discount_percent / 100).quantize(Decimal('0.01')) if o.discount_percent else Decimal('0')
            total += sub - disc
        return total

    # Today stats
    today_qs = Order.objects.filter(status='completed', created_at__date=today).prefetch_related('items')
    today_revenue       = _revenue_from_items(today_qs)
    today_orders_count  = Order.objects.filter(created_at__date=today).count()
    today_completed     = today_qs.count()

    # Weekly stats
    week_qs      = Order.objects.filter(status='completed', created_at__date__gte=week_start).prefetch_related('items')
    week_revenue = _revenue_from_items(week_qs)
    week_orders  = week_qs.count()

    # Monthly stats
    month_qs      = Order.objects.filter(status='completed', created_at__date__gte=month_start).prefetch_related('items')
    month_revenue = _revenue_from_items(month_qs)
    month_orders  = month_qs.count()

    # All-time stats
    all_qs        = Order.objects.filter(status='completed').prefetch_related('items')
    all_revenue   = _revenue_from_items(all_qs)
    all_orders    = all_qs.count()
    avg_order_val = (all_revenue / all_orders) if all_orders else Decimal('0')

    # Top 8 selling items (by quantity sold this month)
    # F is imported at top as F_db
    top_items = (
        OrderItem.objects
        .filter(order__status='completed', order__created_at__date__gte=month_start)
        .values(item_name=F_db('menu_item__name'))
        .annotate(
            total_qty=Sum('quantity'),
            total_rev=Sum(
                ExpressionWrapper(
                    F_db('unit_price') * F_db('quantity'),
                    output_field=DecimalField()
                )
            )
        )
        .order_by('-total_qty')[:8]
    )

    # Last 7 days revenue for chart
    from datetime import timedelta
    chart_labels = []
    chart_values = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        day_orders = Order.objects.filter(status='completed', created_at__date=d).prefetch_related('items')
        rev = _revenue_from_items(day_orders)
        chart_labels.append(d.strftime('%d %b'))
        chart_values.append(float(rev))

    shop = get_shop()
    discounts = Discount.objects.filter(is_active=True)
    pending_ids = [o.id for o in pending_orders]

    return render(request, 'restaurant/staff_portal.html', {
        'tables': tables,
        'pending_orders': pending_orders,
        'accepted_orders': accepted_orders,
        'cooking_orders': cooking_orders,
        'ready_orders': ready_orders,
        'completed_orders': completed_orders,
        'today_revenue': today_revenue,
        'today_orders_count': today_orders_count,
        'today_completed': today_completed,
        'week_revenue': week_revenue,
        'week_orders': week_orders,
        'month_revenue': month_revenue,
        'month_orders': month_orders,
        'all_revenue': all_revenue,
        'all_orders': all_orders,
        'avg_order_val': avg_order_val,
        'top_items': top_items,
        'chart_labels_json': json.dumps(chart_labels),
        'chart_values_json': json.dumps(chart_values),
        'shop': shop,
        'discounts': discounts,
        'pending_ids_json': json.dumps(pending_ids),
    })


@login_required(login_url='/staff/login/')
@require_POST
def update_order_status(request, order_id):
    try:
        order = get_object_or_404(Order, id=order_id)
        data = json.loads(request.body)
        new_status = data.get('status')
        payment_method = data.get('payment_method', '')
        discount_percent = data.get('discount_percent')

        if discount_percent is not None:
            order.discount_percent = Decimal(str(discount_percent))

        valid_transitions = {
            'pending':   ['accepted', 'cancelled'],
            'accepted':  ['preparing', 'cancelled'],
            'preparing': ['ready'],
            'ready':     ['completed'],
            'completed': [], 'cancelled': [],
        }
        if new_status not in valid_transitions.get(order.status, []):
            return JsonResponse({'success': False, 'error': f'Cannot move from {order.status} to {new_status}'})

        order.status = new_status
        if new_status == 'completed':
            order.completed_at = timezone.now()
            if payment_method == 'offline':
                order.payment_status = 'paid_offline'
                order.payment_method = 'offline'
            elif payment_method == 'online':
                order.payment_status = 'paid_online'
                order.payment_method = 'online'
        order.calculate_totals()
        # Persist totals directly so they are never overwritten by model.save() re-calc
        Order.objects.filter(pk=order.pk).update(
            status=order.status,
            completed_at=order.completed_at,
            payment_status=order.payment_status,
            payment_method=order.payment_method,
            discount_percent=order.discount_percent,
            subtotal=order.subtotal,
            discount_amount=order.discount_amount,
            total_amount=order.total_amount,
        )
        # Also call save() so table status logic runs (skip_recalc=True avoids overwriting our update above)
        order.save(skip_recalc=True)
        # When marking ready, tell frontend if there are reorders (for combined bill)
        extra = {}
        if new_status == 'ready':
            root_order = order
            while root_order.parent_order:
                root_order = root_order.parent_order
            reorder_count = root_order.reorders.count()
            extra['has_reorders'] = (order.parent_order is not None) or (reorder_count > 0)
            extra['root_id'] = root_order.id
        return JsonResponse({'success': True, 'status': order.status, **extra})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def pending_orders_api(request):
    """Polling endpoint — returns new pending orders since last_id"""
    try:
        last_id = int(request.GET.get('last_id', 0))
        orders = Order.objects.filter(id__gt=last_id, status='pending').select_related('table')
        result = []
        for o in orders:
            result.append({
                'id': o.id,
                'order_number': o.order_number,
                'customer_name': o.customer_name,
                'table': o.table.name if o.table else 'Takeaway',
                'total': float(o.total_amount),
                'order_type': o.order_type,
                'special_instructions': o.special_instructions or '',
            })
        return JsonResponse({'orders': result})
    except Exception as e:
        return JsonResponse({'orders': [], 'error': str(e)})


def live_pos_orders_api(request):
    """Returns active POS cart snapshots (live, being built) + placed pending/accepted/preparing
    orders — grouped by table — for the staff portal live sidebar.
    """
    try:
        import json as _json
        LIVE_MARKER = '__LIVE__'
        result = []

        # ── 1. Live POS carts (items being added right now in POS terminal) ──
        live_drafts = PosDraft.objects.filter(
            note__startswith=LIVE_MARKER,
            is_deleted=False,
        ).order_by('created_at')

        for d in live_drafts:
            try:
                items = _json.loads(d.items_json or '[]')
            except Exception:
                items = []
            if not items:
                continue   # skip empty carts
            items_preview = ', '.join(
                f"{i.get('qty',1)}× {i.get('name','?')}" for i in items[:3]
            )
            result.append({
                'id': f'draft-{d.id}',
                'order_number': d.draft_number,
                'table': d.table_name or 'POS',
                'order_type': 'takeaway' if (d.table_name or '').startswith('Takeaway') else 'dine_in',
                'status': 'building',          # special status for live carts
                'total': float(d.total_amount),
                'items_preview': items_preview,
                'item_count': len(items),
            })

        # ── 2. Placed orders still active in the kitchen ──
        active_orders = Order.objects.filter(
            status__in=['pending', 'accepted', 'preparing']
        ).select_related('table').prefetch_related('items__menu_item').order_by('created_at')

        for o in active_orders:
            table_label = o.table.name if o.table else ('Takeaway' if o.order_type == 'takeaway' else 'Dine-in')
            items_preview = ', '.join(
                f"{i.quantity}× {i.menu_item.name}" for i in list(o.items.all())[:3]
            )
            result.append({
                'id': o.id,
                'order_number': o.order_number,
                'table': table_label,
                'order_type': o.order_type,
                'status': o.status,
                'total': float(o.total_amount),
                'items_preview': items_preview,
                'item_count': o.items.count(),
            })

        return JsonResponse({'orders': result})
    except Exception as e:
        return JsonResponse({'orders': [], 'error': str(e)})


# ─── EXCEL EXPORT ─────────────────────────────────────────────────


@login_required(login_url='/staff/login/')
def pos_portal(request):
    """POS Table Selection page — entry point for POS at /pos/."""
    shop = get_shop()
    return render(request, 'restaurant/pos_table_selection.html', {
        'shop':        shop,
        'table_range': range(1, 16),   # Order 1–15
        'tw_range':    range(1, 6),    # Takeaway 1–5
    })


@login_required(login_url='/staff/login/')
def pos_terminal(request):
    """POS Terminal — full order entry UI at /pos/terminal/."""
    tables     = Table.objects.filter(is_active=True).order_by('number')
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts  = Discount.objects.filter(is_active=True)
    combos     = Combo.objects.filter(is_active=True).prefetch_related('combo_items__menu_item')
    shop       = get_shop()

    # Build shop JSON for JS bill printing
    import json as _json
    shop_json = _json.dumps({
        'name':    shop.shop_name    if shop else 'Brothers Cafe',
        'address': shop.address      if shop else '',
        'phone':   shop.phone        if shop else '',
        'gstin':   shop.gstin        if shop else '',
        'fssai':   shop.fssai_number if shop else '',
        'upi':     shop.upi_id       if shop else '',
    })

    default_parcel = float(shop.default_parcel_charge) if shop and shop.default_parcel_charge else 0

    today       = timezone.now().date()
    week_start  = today - timezone.timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    def _rev(order_qs):
        total = Decimal('0')
        for o in order_qs.prefetch_related('items'):
            sub  = sum(i.unit_price * i.quantity for i in o.items.all())
            disc = (sub * o.discount_percent / 100).quantize(Decimal('0.01')) if o.discount_percent else Decimal('0')
            total += sub - disc
        return total

    today_qs    = Order.objects.filter(status='completed', created_at__date=today)
    week_qs     = Order.objects.filter(status='completed', created_at__date__gte=week_start)
    month_qs    = Order.objects.filter(status='completed', created_at__date__gte=month_start)
    all_qs      = Order.objects.filter(status='completed')

    today_revenue   = _rev(today_qs)
    today_completed = today_qs.count()
    week_revenue    = _rev(week_qs)
    month_revenue   = _rev(month_qs)
    all_revenue     = _rev(all_qs)
    all_cnt         = all_qs.count()
    avg_order_val   = (all_revenue / all_cnt) if all_cnt else Decimal('0')

    top_items = (
        OrderItem.objects
        .filter(order__status='completed', order__created_at__date__gte=month_start)
        .values(item_name=F_db('menu_item__name'))
        .annotate(
            total_qty=Sum('quantity'),
            total_rev=Sum(ExpressionWrapper(F_db('unit_price') * F_db('quantity'), output_field=DecimalField()))
        )
        .order_by('-total_qty')[:8]
    )

    from datetime import timedelta
    chart_labels, chart_values = [], []
    for i in range(6, -1, -1):
        d  = today - timedelta(days=i)
        qs = Order.objects.filter(status='completed', created_at__date=d)
        chart_labels.append(d.strftime('%d %b'))
        chart_values.append(float(_rev(qs)))

    return render(request, 'restaurant/pos_terminal.html', {
        'tables':               tables,
        'categories':           categories,
        'discounts':            discounts,
        'combos':               combos,
        'shop':                 shop,
        'shop_json':            shop_json,
        'default_parcel':       default_parcel,
        'default_parcel_charge': default_parcel,
        'order_slots':          range(1, 16),   # Order 1–15
        'takeaway_slots':       range(1, 6),    # Takeaway 1–5
        'today_revenue':        today_revenue,
        'today_completed':      today_completed,
        'week_revenue':         week_revenue,
        'month_revenue':        month_revenue,
        'all_revenue':          all_revenue,
        'avg_order_val':        avg_order_val,
        'top_items':            top_items,
        'chart_labels_json':    json.dumps(chart_labels),
        'chart_values_json':    json.dumps(chart_values),
    })

@login_required(login_url='/staff/login/')
def export_excel(request):
    try:
        orders = Order.objects.all().select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        wb = _build_excel(orders, 'All Orders')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="brothers_cafe_all_orders.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)


@login_required(login_url='/staff/login/')
def export_daily_excel(request):
    try:
        import json as _json
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO as _BytesIO
        today = timezone.now().date()
        orders = Order.objects.filter(created_at__date=today).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        drafts = PosDraft.objects.filter(is_deleted=False, created_at__date=today).order_by('-created_at')

        wb = _build_excel(orders, 'Orders')

        # Sheet 2: POS Draft Orders
        ws2 = wb.create_sheet("POS Saved Orders")
        hdr_fill = PatternFill("solid", fgColor="1a1a2e")
        hdr_font = Font(color="FFFFFF", bold=True)
        headers = ['Draft #', 'Time', 'Customer', 'Phone', 'Table', 'Items', 'Subtotal', 'Discount%', 'Total', 'Note']
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for row, d in enumerate(drafts, 2):
            items_str = ', '.join([f"{i.get('qty',1)}x {i.get('name','?')}" for i in _json.loads(d.items_json)])
            ws2.append([d.draft_number, d.created_at.strftime('%H:%M'), d.customer_name,
                        d.customer_phone, d.table_name, items_str,
                        float(d.subtotal), float(d.discount_pct), float(d.total_amount), d.note])
        if drafts.exists():
            ws2.append([])
            ws2.append(['','','','','','TOTAL','',
                        '', float(sum(d.total_amount for d in drafts))])

        buf = _BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_{today}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)






@login_required(login_url='/staff/login/')
def download_qr_codes(request):
    import io, re as _re

    base_url = request.build_absolute_uri('/').rstrip('/')

    tables_data = [
        {"label": t.name, "url": f"{base_url}/table/{t.id}/menu/", "takeaway": False}
        for t in Table.objects.filter(is_active=True).order_by('number')
    ]
    tables_data.append({"label": "Takeaway", "url": f"{base_url}/takeaway/", "takeaway": True})

    def make_qr_svg(url, size=210):
        """Generate QR code: tries segno (pure-python), then qrcode lib, then API image."""
        # 1. Try segno (pure Python, no C deps)
        try:
            import segno
            qr = segno.make_qr(url, error='M')
            buf = io.BytesIO()
            qr.save(buf, kind='svg', scale=4, border=2)
            svg = buf.getvalue().decode('utf-8')
            svg = _re.sub(r'<\?xml[^>]+\?>', '', svg)
            svg = _re.sub(r'<!DOCTYPE[^>]+>', '', svg)
            svg = _re.sub(r'width="[^"]+"', f'width="{size}"', svg, count=1)
            svg = _re.sub(r'height="[^"]+"', f'height="{size}"', svg, count=1)
            return svg.strip()
        except ImportError:
            pass
        except Exception:
            pass
        # 2. Try qrcode library
        try:
            import qrcode
            import qrcode.image.svg as qr_svg
            factory = qr_svg.SvgPathImage
            img = qrcode.make(url, image_factory=factory, box_size=10, border=2)
            buf = io.BytesIO()
            img.save(buf)
            svg = buf.getvalue().decode('utf-8')
            svg = _re.sub(r'<\?xml[^>]+\?>', '', svg)
            svg = _re.sub(r'<!DOCTYPE[^>]+>', '', svg)
            svg = _re.sub(r'width="[^"]+"', f'width="{size}px"', svg, count=1)
            svg = _re.sub(r'height="[^"]+"', f'height="{size}px"', svg, count=1)
            return svg.strip()
        except ImportError:
            pass
        except Exception:
            pass
        # 3. Fallback: use online QR API (works without any library)
        import urllib.parse as _up
        enc = _up.quote(url, safe='')
        return f'<img src="https://api.qrserver.com/v1/create-qr-code/?size={size}x{size}&data={enc}" width="{size}" height="{size}" style="display:block;border-radius:4px" alt="QR Code"/>'

    cards = []
    for t in tables_data:
        svg    = make_qr_svg(t["url"])
        color  = "#e67e22" if t["takeaway"] else "#1a1a2e"
        accent = "#fff8f0" if t["takeaway"] else "#f4f6ff"
        icon   = "🛍️" if t["takeaway"] else "🍽️"
        sub    = "Scan to order takeaway" if t["takeaway"] else "Scan to order food"
        cards.append(f"""<div class="qr-card">
<div class="card-top" style="background:{color}">
  <div class="card-icon">{icon}</div>
  <div><div class="card-name">Brothers Cafe</div><div class="card-sub">{sub}</div></div>
</div>
<div class="card-qr" style="background:{accent}">{svg}</div>
<div class="card-bottom" style="background:{color}">
  <span class="scan-text">☕ Scan &amp; Order</span>
  <span class="staff-label">Staff: {t['label']}</span>
</div></div>""")

    html = '''<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Brothers Cafe QR Codes</title>
<style>
@page{margin:15mm}
@media print{.no-print{display:none!important}body{background:white;padding:0}}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#eef0f3;padding:28px 20px}
h1{text-align:center;font-size:24px;font-weight:800;color:#1a1a2e;margin-bottom:6px}
.sub{text-align:center;font-size:13px;color:#666;margin-bottom:26px}
.grid{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;max-width:820px;margin:0 auto}
.qr-card{background:white;border-radius:16px;overflow:hidden;box-shadow:0 4px 18px rgba(0,0,0,.13);break-inside:avoid}
.card-top{padding:14px 16px;display:flex;align-items:center;gap:10px;color:white}
.card-icon{font-size:26px}.card-name{font-size:14px;font-weight:800}.card-sub{font-size:11px;opacity:.75;margin-top:2px}
.card-qr{padding:16px;display:flex;align-items:center;justify-content:center}
.card-bottom{padding:10px 16px;color:white;display:flex;align-items:center;justify-content:space-between}
.scan-text{font-size:13px;font-weight:700}
.staff-label{font-size:10px;background:rgba(255,255,255,.18);padding:3px 9px;border-radius:20px}
.print-btn{display:block;margin:24px auto 0;padding:13px 38px;background:linear-gradient(135deg,#1a1a2e,#0f3460);color:white;border:none;border-radius:50px;font-size:15px;font-weight:700;cursor:pointer;font-family:inherit}
</style></head><body>
<h1>☕ Brothers Cafe — QR Codes</h1>
<p class="sub">Print &amp; place on each table. Customers scan to order — table numbers never shown to customers.</p>
<div class="grid">''' + ''.join(cards) + '''</div>
<button class="print-btn no-print" onclick="window.print()">🖨️ Print All QR Codes</button>
</body></html>'''

    return HttpResponse(html, content_type='text/html; charset=utf-8')

@login_required(login_url='/staff/login/')
def export_weekly_excel(request):
    try:
        today = timezone.now().date()
        week_start = today - timezone.timedelta(days=today.weekday())
        orders = Order.objects.filter(created_at__date__gte=week_start).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        wb = _build_excel(orders, f'Weekly {week_start} to {today}')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_weekly_{week_start}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)


@login_required(login_url='/staff/login/')
def export_monthly_excel(request):
    try:
        today = timezone.now().date()
        month_start = today.replace(day=1)
        orders = Order.objects.filter(created_at__date__gte=month_start).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        wb = _build_excel(orders, f'Monthly {today.strftime("%B %Y")}')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_{today.strftime("%Y_%m")}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)


@login_required(login_url='/staff/login/')
def export_range_excel(request):
    try:
        from datetime import datetime
        date_from_str = request.GET.get('from', '')
        date_to_str = request.GET.get('to', '')
        if not date_from_str or not date_to_str:
            return HttpResponse('Missing from/to parameters', status=400)
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        orders = Order.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        title = f'{date_from.strftime("%d %b")} to {date_to.strftime("%d %b %Y")}'
        wb = _build_excel(orders, title)
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_{date_from_str}_to_{date_to_str}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)

def _build_excel(orders, title):
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = f'Brothers Cafe — {title}'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='1a1a2e')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Order #','Type','Date','Time','Customer','Phone','Table',
               'Items','Subtotal (Rs)','Disc%','Disc (Rs)','Total (Rs)','Status','Payment']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='e74c3c')
        c.alignment = Alignment(horizontal='center')

    row = 3
    total_rev = Decimal('0')
    for order in orders:
        order_items = list(order.items.select_related('menu_item').all())
        items_str = ', '.join(f"{i.quantity}x {i.menu_item.name}" for i in order_items)
        ldt = timezone.localtime(order.created_at)

        # Recalculate from items in case DB has stale 0 values
        real_subtotal = sum(i.unit_price * i.quantity for i in order_items)
        disc_pct = order.discount_percent or Decimal('0')
        real_disc = (real_subtotal * disc_pct / 100).quantize(Decimal('0.01')) if disc_pct else Decimal('0')
        real_total = real_subtotal - real_disc

        ws.cell(row=row, column=1,  value=order.order_number)
        ws.cell(row=row, column=2,  value=order.get_order_type_display())
        ws.cell(row=row, column=3,  value=ldt.strftime('%d-%m-%Y'))
        ws.cell(row=row, column=4,  value=ldt.strftime('%H:%M'))
        ws.cell(row=row, column=5,  value=order.customer_name)
        ws.cell(row=row, column=6,  value=order.customer_phone)
        ws.cell(row=row, column=7,  value=str(order.table) if order.table else 'Takeaway')
        ws.cell(row=row, column=8,  value=items_str)
        ws.cell(row=row, column=9,  value=float(real_subtotal))
        ws.cell(row=row, column=10, value=float(disc_pct))
        ws.cell(row=row, column=11, value=float(real_disc))
        ws.cell(row=row, column=12, value=float(real_total))
        ws.cell(row=row, column=13, value=order.get_status_display())
        ws.cell(row=row, column=14, value=order.get_payment_status_display())
        if order.status == 'completed':
            total_rev += real_total
        row += 1

    row += 1
    ws.cell(row=row, column=11, value='Total Revenue:').font = Font(bold=True)
    ws.cell(row=row, column=12, value=float(total_rev)).font = Font(bold=True)

    # Set column widths using letters directly (avoids MergedCell.column_letter error)
    col_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']
    for letter in col_letters:
        ws.column_dimensions[letter].width = 16
    ws.column_dimensions['H'].width = 45  # Items column wider
    return wb


# ═══════════════════════════════════════════════════════════════════
# ITEMS PORTAL — Zomato/Swiggy-style menu management for admins
# ═══════════════════════════════════════════════════════════════════

@login_required(login_url='/staff/login/')
def items_portal(request):
    """Admin-facing items management portal grouped by category."""
    categories = Category.objects.prefetch_related('items').order_by('order', 'name')
    shop = get_shop()
    return render(request, 'restaurant/items_portal.html', {
        'categories': categories,
        'shop': shop,
    })


@login_required(login_url='/staff/login/')
def items_portal_toggle(request, item_id):
    """Toggle dine-in / takeaway / featured for an item via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        import json as _json
        data = _json.loads(request.body)
        field = data.get('field')
        item = MenuItem.objects.get(pk=item_id)
        if field == 'dine_in':
            item.is_available_dine_in = not item.is_available_dine_in
            item.save(update_fields=['is_available_dine_in'])
            return JsonResponse({'success': True, 'value': item.is_available_dine_in})
        elif field == 'takeaway':
            item.is_available_takeaway = not item.is_available_takeaway
            item.save(update_fields=['is_available_takeaway'])
            return JsonResponse({'success': True, 'value': item.is_available_takeaway})
        elif field == 'featured':
            item.is_featured = not item.is_featured
            item.save(update_fields=['is_featured'])
            return JsonResponse({'success': True, 'value': item.is_featured})
        return JsonResponse({'error': 'Unknown field'}, status=400)
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def items_portal_update_price(request, item_id):
    """Quick-update price for a menu item via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        import json as _json
        data = _json.loads(request.body)
        price = Decimal(str(data.get('price', 0)))
        if price < 0:
            return JsonResponse({'error': 'Price must be non-negative'}, status=400)
        item = MenuItem.objects.get(pk=item_id)
        item.price = price
        item.save(update_fields=['price'])
        return JsonResponse({'success': True, 'price': float(price)})
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def items_portal_delete_item(request, item_id):
    """Delete a menu item via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        item = MenuItem.objects.get(pk=item_id)
        item.delete()
        return JsonResponse({'success': True})
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ═══════════════════════════════════════════════════════════════════
# ITEMS PORTAL — Add / Edit / Upload image — inline (no admin redirect)
# ═══════════════════════════════════════════════════════════════════

@login_required(login_url='/staff/login/')
def items_portal_add_item(request):
    """Add a new menu item via AJAX (supports image upload)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        name         = request.POST.get('name', '').strip()
        category_id  = request.POST.get('category_id', '')
        price        = request.POST.get('price', '0')
        parcel_charge= request.POST.get('parcel_charge', '0')
        item_type    = request.POST.get('item_type', 'veg')
        description  = request.POST.get('description', '').strip()
        dine_in      = request.POST.get('is_available_dine_in', 'true') == 'true'
        takeaway     = request.POST.get('is_available_takeaway', 'true') == 'true'
        featured     = request.POST.get('is_featured', 'false') == 'true'

        if not name:
            return JsonResponse({'error': 'Item name is required'}, status=400)
        category = Category.objects.get(pk=category_id)

        item = MenuItem(
            name=name,
            category=category,
            price=Decimal(price),
            parcel_charge=Decimal(parcel_charge),
            item_type=item_type,
            description=description,
            is_available_dine_in=dine_in,
            is_available_takeaway=takeaway,
            is_featured=featured,
            is_available=True,
        )
        if 'image' in request.FILES:
            item.image = request.FILES['image']
        item.save()

        return JsonResponse({
            'success': True,
            'item_id': item.id,
            'name': item.name,
            'price': float(item.price),
            'image_url': item.image.url if item.image else None,
            'item_type': item.item_type,
        })
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def items_portal_edit_item(request, item_id):
    """Edit a menu item via AJAX (supports image upload)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        item = MenuItem.objects.get(pk=item_id)

        name         = request.POST.get('name', item.name).strip()
        category_id  = request.POST.get('category_id', item.category_id)
        price        = request.POST.get('price', str(item.price))
        parcel_charge= request.POST.get('parcel_charge', str(item.parcel_charge))
        item_type    = request.POST.get('item_type', item.item_type)
        description  = request.POST.get('description', item.description).strip()
        dine_in      = request.POST.get('is_available_dine_in', 'true') == 'true'
        takeaway     = request.POST.get('is_available_takeaway', 'true') == 'true'
        featured     = request.POST.get('is_featured', 'false') == 'true'

        item.name          = name
        item.category_id   = category_id
        item.price         = Decimal(price)
        item.parcel_charge = Decimal(parcel_charge)
        item.item_type     = item_type
        item.description   = description
        item.is_available_dine_in   = dine_in
        item.is_available_takeaway  = takeaway
        item.is_featured            = featured
        if 'image' in request.FILES:
            item.image = request.FILES['image']
        item.save()

        return JsonResponse({
            'success': True,
            'name': item.name,
            'price': float(item.price),
            'image_url': item.image.url if item.image else None,
        })
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def items_portal_get_item(request, item_id):
    """Return item data as JSON for the edit modal."""
    try:
        item = MenuItem.objects.select_related('category').get(pk=item_id)
        return JsonResponse({
            'id': item.id,
            'name': item.name,
            'category_id': item.category_id,
            'price': float(item.price),
            'parcel_charge': float(item.parcel_charge),
            'item_type': item.item_type,
            'description': item.description,
            'is_available_dine_in': item.is_available_dine_in,
            'is_available_takeaway': item.is_available_takeaway,
            'is_featured': item.is_featured,
            'image_url': item.image.url if item.image else None,
        })
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


# ═══════════════════════════════════════════════════════════════════
# MENU MANAGER — Full inline CRUD APIs (no Django admin)
# ═══════════════════════════════════════════════════════════════════

# ── CATEGORIES ──────────────────────────────────────────────────
@login_required(login_url='/staff/login/')
def mm_categories(request):
    cats = list(Category.objects.values('id','name','icon','description','order','is_active'))
    for c in cats:
        c['item_count'] = MenuItem.objects.filter(category_id=c['id']).count()
    return JsonResponse({'categories': cats})

@login_required(login_url='/staff/login/')
def mm_category_save(request, cat_id=None):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        data = json.loads(request.body)
        if cat_id:
            cat = Category.objects.get(pk=cat_id)
        else:
            cat = Category()
        cat.name = data.get('name','').strip()
        cat.icon = data.get('icon','').strip()
        cat.description = data.get('description','').strip()
        cat.order = int(data.get('order', cat.order if cat_id else 0))
        cat.is_active = data.get('is_active', True)
        cat.save()
        return JsonResponse({'success':True,'id':cat.id,'name':cat.name})
    except Category.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

@login_required(login_url='/staff/login/')
def mm_category_toggle(request, cat_id):
    """Quick toggle is_active for a category without changing other fields."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        cat = Category.objects.get(pk=cat_id)
        cat.is_active = not cat.is_active
        cat.save(update_fields=['is_active'])
        return JsonResponse({'success': True, 'is_active': cat.is_active})
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url='/staff/login/')
def mm_category_delete(request, cat_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        Category.objects.get(pk=cat_id).delete()
        return JsonResponse({'success':True})
    except Category.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

# ── COMBOS ──────────────────────────────────────────────────────
@login_required(login_url='/staff/login/')
def mm_combos(request):
    combos = []
    for c in Combo.objects.prefetch_related('combo_items__menu_item').all():
        combos.append({
            'id':c.id,'name':c.name,'description':c.description,
            'price':float(c.price),'icon':c.icon,'is_active':c.is_active,'order':c.order,
            'item_count':c.combo_items.count(),
            'items':[{'name':ci.menu_item.name,'qty':ci.quantity} for ci in c.combo_items.all()]
        })
    return JsonResponse({'combos':combos})

@login_required(login_url='/staff/login/')
def mm_combo_save(request, combo_id=None):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        data = json.loads(request.body)
        if combo_id:
            combo = Combo.objects.get(pk=combo_id)
        else:
            combo = Combo()
        combo.name = data.get('name','').strip()
        combo.description = data.get('description','').strip()
        combo.price = Decimal(str(data.get('price',0)))
        combo.icon = data.get('icon','🎁').strip()
        combo.is_active = data.get('is_active', True)
        combo.order = int(data.get('order', combo.order if combo_id else 0))
        combo.save()
        return JsonResponse({'success':True,'id':combo.id})
    except Combo.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

@login_required(login_url='/staff/login/')
def mm_combo_toggle(request, combo_id):
    """Quick toggle is_active for a combo without changing other fields."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        combo = Combo.objects.get(pk=combo_id)
        combo.is_active = not combo.is_active
        combo.save(update_fields=['is_active'])
        return JsonResponse({'success': True, 'is_active': combo.is_active})
    except Combo.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url='/staff/login/')
def mm_combo_delete(request, combo_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        Combo.objects.get(pk=combo_id).delete()
        return JsonResponse({'success':True})
    except Combo.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

# ── DISCOUNTS ────────────────────────────────────────────────────
@login_required(login_url='/staff/login/')
def mm_discounts(request):
    discs = list(Discount.objects.values('id','name','percent','is_active','valid_from','valid_to','description'))
    for d in discs:
        d['valid_from'] = str(d['valid_from']) if d['valid_from'] else ''
        d['valid_to'] = str(d['valid_to']) if d['valid_to'] else ''
        d['percent'] = float(d['percent'])
    return JsonResponse({'discounts':discs})

@login_required(login_url='/staff/login/')
def mm_discount_save(request, disc_id=None):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        data = json.loads(request.body)
        if disc_id:
            disc = Discount.objects.get(pk=disc_id)
        else:
            disc = Discount()
        disc.name = data.get('name','').strip()
        disc.percent = Decimal(str(data.get('percent',0)))
        disc.is_active = data.get('is_active', True)
        disc.description = data.get('description','').strip()
        vf = data.get('valid_from','').strip()
        vt = data.get('valid_to','').strip()
        disc.valid_from = vf if vf else None
        disc.valid_to = vt if vt else None
        disc.save()
        return JsonResponse({'success':True,'id':disc.id})
    except Discount.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

@login_required(login_url='/staff/login/')
def mm_discount_delete(request, disc_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        Discount.objects.get(pk=disc_id).delete()
        return JsonResponse({'success':True})
    except Discount.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

# ── TABLES ───────────────────────────────────────────────────────
@login_required(login_url='/staff/login/')
def mm_tables(request):
    tables = list(Table.objects.values('id','number','name','capacity','status','description','is_active'))
    return JsonResponse({'tables':tables})

@login_required(login_url='/staff/login/')
def mm_table_save(request, table_id=None):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        data = json.loads(request.body)
        if table_id:
            tbl = Table.objects.get(pk=table_id)
        else:
            tbl = Table()
        tbl.number = int(data.get('number', tbl.number if table_id else 1))
        tbl.name = data.get('name','').strip()
        tbl.capacity = int(data.get('capacity', 4))
        tbl.description = data.get('description','').strip()
        tbl.is_active = data.get('is_active', True)
        tbl.save()
        return JsonResponse({'success':True,'id':tbl.id})
    except Table.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

@login_required(login_url='/staff/login/')
def mm_table_delete(request, table_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        Table.objects.get(pk=table_id).delete()
        return JsonResponse({'success':True})
    except Table.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

# ── CUSTOMERS ────────────────────────────────────────────────────
@login_required(login_url='/staff/login/')
def mm_customers(request):
    custs = []
    for c in CustomerProfile.objects.all()[:100]:
        custs.append({
            'id':c.id,'name':c.name,'phone':c.phone,
            'visit_count':c.visit_count,
            'total_orders':c.total_orders,
            'total_spent':float(c.total_spent),
            'last_visit':c.last_visit.strftime('%d %b %Y'),
        })
    return JsonResponse({'customers':custs})

# ── SHOP SETTINGS ────────────────────────────────────────────────
@login_required(login_url='/staff/login/')
def mm_shop_settings(request):
    shop = ShopSettings.objects.first()
    if not shop: return JsonResponse({'settings':None})
    return JsonResponse({'settings':{
        'id':shop.id,'shop_name':shop.shop_name,'location':shop.location,
        'gstin':shop.gstin,'fssai_number':shop.fssai_number,
        'phone':shop.phone,'email':shop.email,'address':shop.address,
        'upi_id':shop.upi_id,
        'default_discount_percent':float(shop.default_discount_percent),
        'default_parcel_charge':float(shop.default_parcel_charge),
    }})

@login_required(login_url='/staff/login/')
def mm_shop_settings_save(request):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        data = json.loads(request.body)
        shop = ShopSettings.objects.first()
        if not shop: shop = ShopSettings()
        shop.shop_name = data.get('shop_name','Brothers Cafe').strip()
        shop.location = data.get('location','').strip()
        shop.gstin = data.get('gstin','').strip()
        shop.fssai_number = data.get('fssai_number','').strip()
        shop.phone = data.get('phone','').strip()
        shop.email = data.get('email','').strip()
        shop.address = data.get('address','').strip()
        shop.upi_id = data.get('upi_id','').strip()
        shop.default_discount_percent = Decimal(str(data.get('default_discount_percent',0)))
        shop.default_parcel_charge = Decimal(str(data.get('default_parcel_charge',0)))
        shop.save()
        return JsonResponse({'success':True})
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

# ── ORDERS LIST ──────────────────────────────────────────────────
@login_required(login_url='/staff/login/')
def mm_orders(request):
    page = int(request.GET.get('page',1))
    status = request.GET.get('status','')
    qs = Order.objects.select_related('table').order_by('-created_at')
    if status: qs = qs.filter(status=status)
    from django.core.paginator import Paginator
    p = Paginator(qs, 20)
    pg = p.get_page(page)
    orders = []
    for o in pg:
        orders.append({
            'id':o.id,'order_number':o.order_number,'customer_name':o.customer_name,
            'customer_phone':o.customer_phone,'table':str(o.table) if o.table else 'Takeaway',
            'status':o.status,'order_type':o.order_type,
            'total_amount':float(o.total_amount),
            'created_at':o.created_at.strftime('%d %b %Y %H:%M'),
            'payment_status':o.payment_status,
        })
    return JsonResponse({'orders':orders,'total_pages':p.num_pages,'page':page})

# ── MENU MANAGER MAIN PAGE ────────────────────────────────────────
@login_required(login_url='/staff/login/')
def menu_manager(request):
    """Main menu manager page — renders the full management UI."""
    categories = Category.objects.prefetch_related('items').order_by('order','name')
    menu_items = MenuItem.objects.select_related('category').order_by('category__name','name')
    all_categories = Category.objects.order_by('order','name')
    combos = Combo.objects.prefetch_related('combo_items__menu_item').all()
    discounts = Discount.objects.all()
    tables = Table.objects.order_by('number')
    customers = CustomerProfile.objects.all()[:50]
    shop = ShopSettings.objects.first()

    # Stats
    total_items = MenuItem.objects.count()
    available_items = MenuItem.objects.filter(is_available_dine_in=True).count()
    avg_price = MenuItem.objects.aggregate(a=models_Avg('price'))['a'] or 0

    return render(request, 'restaurant/menu_manager.html', {
        'categories': categories,
        'menu_items': menu_items,
        'all_categories': all_categories,
        'combos': combos,
        'discounts': discounts,
        'tables': tables,
        'customers': customers,
        'shop': shop,
        'total_items': total_items,
        'available_items': available_items,
        'avg_price': avg_price,
        'cat_count': all_categories.count(),
    })
