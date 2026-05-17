from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils import timezone
from django.db.models import Sum, Prefetch

# ── GF(256) lookup tables for QR code generation ─────────────────────────────
GF256_EXP = [0] * 512
GF256_LOG  = [0] * 256
_gfx = 1
for _gfi in range(255):
    GF256_EXP[_gfi] = _gfx
    GF256_LOG[_gfx] = _gfi
    _gfx <<= 1
    if _gfx & 0x100: _gfx ^= 0x11d
for _gfi in range(255, 512): GF256_EXP[_gfi] = GF256_EXP[_gfi - 255]
# ─────────────────────────────────────────────────────────────────────────────
from .models import Table, Category, MenuItem, Order, OrderItem, ShopSettings, Discount, CustomerProfile, Combo, ComboItem
import json
import openpyxl
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP


# ═══════════════════════════════════════════════════════════════════
# CUSTOMER LOGIN & HISTORY
# ═══════════════════════════════════════════════════════════════════

def get_shop():
    return ShopSettings.objects.first()

def customer_login(request):
    """
    Combined Register / Login page.
    Register: name + phone → creates CustomerProfile, sets session.
    Login: phone only → looks up existing profile, sets session.
    """
    next_url = request.GET.get('next', '/')
    shop = get_shop()

    # Already logged in → go to home
    if request.session.get('customer_phone'):
        return redirect('/customer/home/')

    context = {'next': next_url, 'shop': shop}

    if request.method == 'POST':
        action = request.POST.get('action', 'register')
        next_url = request.POST.get('next', '/')

        if action == 'register':
            name  = request.POST.get('name', '').strip()
            phone = request.POST.get('phone', '').strip()
            context['reg_name']  = name
            context['reg_phone'] = phone

            if not name:
                context['reg_error'] = 'Please enter your name.'
            elif not phone or len(phone) < 8:
                context['reg_error'] = 'Please enter a valid phone number.'
            elif CustomerProfile.objects.filter(phone=phone).exists():
                context['reg_error'] = 'This phone number is already registered. Please use Login.'
            else:
                profile = CustomerProfile.objects.create(name=name, phone=phone)
                request.session['customer_phone'] = phone
                request.session['customer_name']  = name
                request.session['customer_id']    = profile.id
                return redirect(next_url)

        elif action == 'login':
            phone = request.POST.get('phone', '').strip()
            context['login_phone'] = phone

            if not phone or len(phone) < 8:
                context['login_error'] = 'Please enter a valid phone number.'
            else:
                try:
                    profile = CustomerProfile.objects.get(phone=phone)
                    profile.visit_count += 1
                    profile.save()
                    request.session['customer_phone'] = phone
                    request.session['customer_name']  = profile.name
                    request.session['customer_id']    = profile.id
                    return redirect('/customer/home/')
                except CustomerProfile.DoesNotExist:
                    context['login_error'] = 'Phone number not found. Please register first.'

    return render(request, 'restaurant/customer_login.html', context)


def customer_logout(request):
    request.session.flush()
    return redirect('/')


def customer_history(request):
    phone = request.session.get('customer_phone')
    if not phone:
        return redirect("/customer/login/?next=/customer/history/")
    shop = get_shop()
    name = request.session.get('customer_name', '')

    all_orders = Order.objects.filter(
        customer_phone=phone,
        parent_order__isnull=True,
    ).prefetch_related('items__menu_item','reorders__items__menu_item').order_by('-created_at')

    def enrich(order):
        linked = [order] + list(order.reorders.order_by('created_at'))
        combined_total = sum(o.total_amount for o in linked)
        all_items = []
        for o in linked:
            for item in o.items.select_related('menu_item').all():
                item.order = o
                all_items.append(item)
        return {
            'order': order,
            'combined_total': combined_total,
            'all_items': all_items,
            'all_orders': linked,      # for live card to show per-order grouping
            'reorder_count': len(linked)-1,
        }

    # Active = any non-completed/non-cancelled root order
    ACTIVE_STATUSES = ['pending','accepted','preparing','ready']
    active_entry = None
    past_entries = []
    for o in all_orders:
        e = enrich(o)
        if o.status in ACTIVE_STATUSES and active_entry is None:
            active_entry = e
        else:
            past_entries.append(e)

    all_entries = [enrich(o) for o in all_orders]

    # Stats
    completed_count = sum(1 for o in all_orders if o.status == 'completed')
    from django.db.models import Sum
    total_spent = Order.objects.filter(
        customer_phone=phone, status='completed', parent_order__isnull=True
    ).aggregate(t=Sum('total_amount'))['t'] or 0

    # Last table used for "Place New Order" button
    last_order = all_orders.first()
    last_table_id = last_order.table.id if last_order and last_order.table else None
    is_takeaway = last_order.order_type == 'takeaway' if last_order else False

    prog_steps = [
        ('pending',   '📋', 'Received'),
        ('accepted',  '✅', 'Accepted'),
        ('preparing', '👨‍🍳', 'Cooking'),
        ('ready',     '🔔', 'Ready'),
        ('completed', '🎉', 'Done'),
    ]

    return render(request, 'restaurant/customer_home.html', {
        'orders': all_entries,
        'active_order': active_entry,
        'past_orders': past_entries,
        'customer_name': name,
        'customer_phone': phone,
        'completed_count': completed_count,
        'total_spent': total_spent,
        'last_table_id': last_table_id,
        'is_takeaway': is_takeaway,
        'prog_steps': prog_steps,
        'shop': shop,
    })


def table_selection(request):
    """Root URL — redirect customers to login, show nothing useful."""
    # If already logged in, show a simple welcome with history link
    if request.session.get('customer_phone'):
        name = request.session.get('customer_name','')
        return redirect('/customer/history/')
    return redirect('/customer/login/?next=/customer/history/')


def _get_active_order_for_menu(phone):
    """Return active order entry for a customer phone, or None."""
    if not phone:
        return None
    ACTIVE_STATUSES = ['pending', 'accepted', 'preparing', 'ready']
    order = Order.objects.filter(
        customer_phone=phone,
        parent_order__isnull=True,
        status__in=ACTIVE_STATUSES,
    ).prefetch_related('items__menu_item', 'reorders__items__menu_item').order_by('-created_at').first()
    if not order:
        return None
    linked = [order] + list(order.reorders.order_by('created_at'))
    combined_total = sum(o.total_amount for o in linked)
    return {'order': order, 'combined_total': combined_total, 'all_orders': linked}


def _get_order_history_for_menu(phone, limit=5):
    """Return past orders for a customer (completed/cancelled), newest first."""
    if not phone:
        return []
    past = Order.objects.filter(
        customer_phone=phone,
        parent_order__isnull=True,
        status__in=['completed', 'cancelled'],
    ).prefetch_related('items__menu_item', 'reorders__items__menu_item').order_by('-created_at')[:limit]
    result = []
    for o in past:
        linked = [o] + list(o.reorders.order_by('created_at'))
        combined_total = sum(x.total_amount for x in linked)
        all_items = []
        for x in linked:
            for item in x.items.select_related('menu_item').all():
                all_items.append(item)
        result.append({'order': o, 'combined_total': combined_total, 'all_items': all_items, 'reorder_count': len(linked) - 1})
    return result


def menu_view(request, table_id):
    # Require customer login (skip for reorders — session already set)
    reorder_from = request.GET.get('reorder_from')
    if not request.session.get('customer_phone') and not reorder_from:
        return redirect(f"/customer/login/?next=/table/{table_id}/menu/")

    table = get_object_or_404(Table, id=table_id, is_active=True)
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts = Discount.objects.filter(is_active=True)
    shop = get_shop()
    reorder_order = None
    reorder_customer_name = ''
    reorder_customer_phone = ''
    if reorder_from:
        try:
            # Walk up to root order so we always link to the original
            reorder_order = Order.objects.get(id=reorder_from)
            while reorder_order.parent_order:
                reorder_order = reorder_order.parent_order
            reorder_customer_name = reorder_order.customer_name
            reorder_customer_phone = reorder_order.customer_phone
        except Order.DoesNotExist:
            reorder_order = None

    phone = request.session.get('customer_phone', '')
    active_order = _get_active_order_for_menu(phone)
    past_orders = _get_order_history_for_menu(phone)
    combos = Combo.objects.filter(is_active=True).prefetch_related('combo_items__menu_item')

    return render(request, 'restaurant/menu.html', {
        'table': table, 'categories': categories,
        'discounts': discounts, 'shop': shop,
        'reorder_from': str(reorder_order.id) if reorder_order else None,
        'reorder_order': reorder_order,
        'is_takeaway': False,
        'active_order': active_order,
        'past_orders': past_orders,
        'combos': combos,
    })


@ensure_csrf_cookie
def takeaway_menu(request):
    """Takeaway order — no table needed"""
    if not request.session.get('customer_phone'):
        return redirect("/customer/login/?next=/takeaway/")
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts = Discount.objects.filter(is_active=True)
    shop = get_shop()
    phone = request.session.get('customer_phone', '')
    active_order = _get_active_order_for_menu(phone)
    past_orders = _get_order_history_for_menu(phone)
    combos = Combo.objects.filter(is_active=True).prefetch_related('combo_items__menu_item')
    return render(request, 'restaurant/menu.html', {
        'table': None, 'categories': categories,
        'discounts': discounts, 'shop': shop,
        'cart': '{}', 'cart_count': 0,
        'is_takeaway': True,
        'active_order': active_order,
        'past_orders': past_orders,
        'combos': combos,
    })


def place_order(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
        customer_name  = (data.get('customer_name') or request.session.get('customer_name', '')).strip()
        customer_phone = (data.get('customer_phone') or request.session.get('customer_phone', '')).strip()
        special_instructions = data.get('special_instructions', '')
        items = data.get('items', [])
        discount_percent = Decimal(str(data.get('discount_percent', 0)))
        order_type = data.get('order_type', 'dine_in')  # 'dine_in' or 'takeaway'
        table_id = data.get('table_id')
        parent_order_id = data.get('parent_order_id')
        reorder_from_id = data.get('reorder_from')  # parent order id for reorders

        if not customer_name:
            return JsonResponse({'success': False, 'error': 'Customer name is required'})
        if not customer_phone:
            return JsonResponse({'success': False, 'error': 'Phone number is required'})
        if not items:
            return JsonResponse({'success': False, 'error': 'No items in order'})

        table = None
        if order_type == 'dine_in' and table_id:
            table = Table.objects.filter(id=table_id, is_active=True).first()
            if not table:
                return JsonResponse({'success': False, 'error': 'Table not found'})

        special_note = special_instructions
        parent_order_obj = None
        if parent_order_id:
            try:
                parent_order_obj = Order.objects.get(id=parent_order_id)
                special_note = f'Reorder (linked to #{parent_order_obj.order_number})' + (f' — {special_instructions}' if special_instructions else '')
            except Order.DoesNotExist:
                pass

        import random
        order = Order(
            table=table,
            customer_name=customer_name,
            customer_phone=customer_phone,
            special_instructions=special_note,
            discount_percent=discount_percent,
            order_type=order_type,
            order_number=f"BC{timezone.now().strftime('%Y%m%d')}{random.randint(1000,9999)}",
            subtotal=Decimal('0'),
            discount_amount=Decimal('0'),
            total_amount=Decimal('0'),
            parent_order=parent_order_obj,
        )
        order.save()

        subtotal = Decimal('0')
        for item_data in items:
            menu_item = MenuItem.objects.filter(id=item_data['id'], is_available=True).first()
            if not menu_item:
                continue
            qty = max(1, int(item_data.get('qty', 1)))
            OrderItem.objects.create(order=order, menu_item=menu_item,
                                     quantity=qty, unit_price=menu_item.price)
            subtotal += menu_item.price * qty

        order.subtotal = subtotal
        if subtotal > 0 and discount_percent > 0:
            order.discount_amount = (subtotal * discount_percent / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            order.discount_amount = Decimal('0.00')
        order.total_amount = order.subtotal - order.discount_amount
        Order.objects.filter(pk=order.pk).update(
            subtotal=order.subtotal,
            discount_amount=order.discount_amount,
            total_amount=order.total_amount,
        )
        if table:
            Table.objects.filter(pk=table.pk).update(status='occupied')

        return JsonResponse({'success': True, 'order_id': order.id, 'order_number': order.order_number, 'history_url': '/customer/history/'})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid request data'}, status=400)
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@ensure_csrf_cookie
def reorder_menu(request, order_id):
    """Show menu so customer can pick new items — these will be added to existing order bill"""
    order = get_object_or_404(Order, id=order_id)
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts = Discount.objects.filter(is_active=True)
    shop = get_shop()
    import json as _json
    prefill = {str(item.menu_item.id): {'name': item.menu_item.name, 'price': float(item.unit_price), 'qty': item.quantity}
               for item in order.items.all()}
    return render(request, 'restaurant/menu.html', {
        'table': order.table,
        'categories': categories,
        'discounts': discounts,
        'shop': shop,
        'cart': _json.dumps(prefill),
        'cart_count': sum(v['qty'] for v in prefill.values()),
        'reorder_for': order,
        'is_reorder': True,
    })


def add_items_to_order(request, order_id):
    """Add new items from reorder menu into the existing order bill"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        order = get_object_or_404(Order, id=order_id)
        data = json.loads(request.body)
        items = data.get('items', [])
        if not items:
            return JsonResponse({'success': False, 'error': 'No items provided'})
        for item_data in items:
            menu_item = MenuItem.objects.filter(id=item_data['id'], is_available=True).first()
            if not menu_item:
                continue
            qty = max(1, int(item_data.get('qty', 1)))
            existing = order.items.filter(menu_item=menu_item).first()
            if existing:
                existing.quantity += qty
                existing.save()
            else:
                OrderItem.objects.create(order=order, menu_item=menu_item,
                                         quantity=qty, unit_price=menu_item.price)
        order.calculate_totals()
        Order.objects.filter(pk=order.pk).update(
            subtotal=order.subtotal,
            discount_amount=order.discount_amount,
            total_amount=order.total_amount,
        )
        return JsonResponse({'success': True, 'order_id': order.id, 'order_number': order.order_number, 'history_url': '/customer/history/'})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def reorder(request, order_id):
    """Redirect customer to menu with original items pre-loaded for reorder"""
    orig = get_object_or_404(Order, id=order_id)
    if orig.table:
        return redirect(f"/table/{orig.table.id}/menu/?reorder_from={orig.id}")
    else:
        return redirect(f"/takeaway/?reorder_from={orig.id}")


def order_status(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    shop = get_shop()

    # Find root order and gather all orders in this session (first + reorders)
    root_order = order
    while root_order.parent_order:
        root_order = root_order.parent_order
    all_orders = [root_order] + list(root_order.reorders.order_by('created_at'))

    # Build combined items list and total across all orders
    combined_items = []
    combined_subtotal = Decimal('0')
    for o in all_orders:
        for item in o.items.select_related('menu_item').all():
            combined_items.append({
                'name': item.menu_item.name,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price,
                'is_reorder': o.parent_order is not None,
            })
            combined_subtotal += item.total_price
    discount_percent = root_order.discount_percent
    discount_amount = (combined_subtotal * discount_percent / 100).quantize(Decimal('0.01')) if discount_percent else Decimal('0')
    combined_total = combined_subtotal - discount_amount

    # Bill is ready only when every order in session is completed
    all_completed = all(o.status == 'completed' for o in all_orders)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': order.status,           # this order's status (for progress bar)
            'root_status': root_order.status, # root order status
            'all_completed': all_completed,   # True only when EVERY order is done
            'payment_status': order.payment_status,
            'status_display': order.get_status_display(),
            'combined_total': float(combined_total),
            'has_reorders': len(all_orders) > 1,
        })
    return render(request, 'restaurant/order_status.html', {
        'order': order,
        'root_order': root_order,
        'all_orders': all_orders,
        'combined_items': combined_items,
        'combined_subtotal': combined_subtotal,
        'discount_percent': discount_percent,
        'discount_amount': discount_amount,
        'combined_total': combined_total,
        'has_reorders': len(all_orders) > 1,
        'all_completed': all_completed,
        'shop': shop,
    })


def bill_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    shop = get_shop()
    # Find root order - bill is only available when root order is completed
    root_order = order
    while root_order.parent_order:
        root_order = root_order.parent_order
    all_orders_check = [root_order] + list(root_order.reorders.order_by('created_at'))
    all_done = all(o.status == 'completed' for o in all_orders_check)
    if not all_done:
        return render(request, 'restaurant/waiting_for_bill.html', {'order': root_order, 'shop': shop})
    all_orders = [root_order] + list(root_order.reorders.order_by('created_at'))
    combined_items = []
    combined_subtotal = Decimal('0')
    for o in all_orders:
        for item in o.items.select_related('menu_item').all():
            combined_items.append({
                'order_number': o.order_number,
                'name': item.menu_item.name,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price,
                'is_reorder': o.parent_order is not None,
            })
            combined_subtotal += item.total_price
    discount_percent = root_order.discount_percent
    discount_amount = (combined_subtotal * discount_percent / 100).quantize(Decimal('0.01')) if discount_percent else Decimal('0')
    combined_total = combined_subtotal - discount_amount
    return render(request, 'restaurant/bill.html', {
        'order': order,
        'root_order': root_order,
        'all_orders': all_orders,
        'combined_items': combined_items,
        'combined_subtotal': combined_subtotal,
        'discount_percent': discount_percent,
        'discount_amount': discount_amount,
        'combined_total': combined_total,
        'has_reorders': len(all_orders) > 1,
        'shop': shop,
    })


def pay_online(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    shop = get_shop()
    order.payment_method = 'online'
    order.payment_status = 'online_pending'
    order.save()
    return render(request, 'restaurant/pay_online.html', {'order': order, 'shop': shop})


def confirm_payment(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if request.method == 'POST':
        order.payment_status = 'paid_online'
        order.payment_method = 'online'
        order.completed_at = timezone.now()
        order.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})


def download_bill(request, order_id):
    """Keep for backward compat but redirect to bill view"""
    return redirect('bill', order_id=order_id)




def print_bill(request, order_id):
    """Thermal printer-friendly bill page for staff"""
    order = get_object_or_404(Order, id=order_id)
    shop = get_shop()
    root_order = order
    while root_order.parent_order:
        root_order = root_order.parent_order
    all_orders = [root_order] + list(root_order.reorders.order_by('created_at'))
    combined_items = []
    combined_subtotal = Decimal('0')
    for o in all_orders:
        for item in o.items.select_related('menu_item').all():
            combined_items.append({
                'name': item.menu_item.name,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price,
            })
            combined_subtotal += item.total_price
    discount_percent = root_order.discount_percent
    discount_amount = (combined_subtotal * discount_percent / 100).quantize(Decimal('0.01')) if discount_percent else Decimal('0')
    combined_total = combined_subtotal - discount_amount
    return render(request, 'restaurant/print_bill.html', {
        'order': order,
        'root_order': root_order,
        'all_orders': all_orders,
        'combined_items': combined_items,
        'combined_subtotal': combined_subtotal,
        'discount_percent': discount_percent,
        'discount_amount': discount_amount,
        'combined_total': combined_total,
        'shop': shop,
    })



def reorder_notification_api(request):
    """Returns new reorder (child) orders since last_id for staff sound alert"""
    try:
        last_id = int(request.GET.get('last_id', 0))
        orders = Order.objects.filter(
            id__gt=last_id,
            parent_order__isnull=False,
            status='pending'
        ).select_related('table', 'parent_order')
        result = []
        for o in orders:
            result.append({
                'id': o.id,
                'order_number': o.order_number,
                'customer_name': o.customer_name,
                'table': o.table.name if o.table else 'Takeaway',
                'total': float(o.total_amount),
                'parent_order_number': o.parent_order.order_number if o.parent_order else '',
                'is_reorder': True,
            })
        return JsonResponse({'orders': result})
    except Exception as e:
        return JsonResponse({'orders': [], 'error': str(e)})

# ─── STAFF PORTAL ─────────────────────────────────────────────────

def staff_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user and user.is_staff:
            login(request, user)
            return redirect('staff_portal')
        return render(request, 'restaurant/staff_login.html', {'error': 'Invalid credentials'})
    return render(request, 'restaurant/staff_login.html')


def staff_logout(request):
    logout(request)
    return redirect('staff_login')



# ═══════════════════════════════════════════════════════════════════
# LIVE ORDER PAGE
# ═══════════════════════════════════════════════════════════════════

def live_order(request):
    """Dedicated live order page — reads order_id from query param or session."""
    if not request.session.get('customer_phone'):
        return redirect('/customer/login/')
    return render(request, 'restaurant/live_order.html', {})


def live_order_data(request):
    """JSON API for the live order page — returns full order state."""
    order_id = request.GET.get('order_id')
    if not order_id:
        return JsonResponse({'error': 'no order id'}, status=400)
    try:
        root = Order.objects.get(id=order_id)
        # Walk to root
        while root.parent_order:
            root = root.parent_order
    except Order.DoesNotExist:
        return JsonResponse({'error': 'not found'}, status=404)

    all_orders = [root] + list(root.reorders.order_by('created_at'))
    all_completed = all(o.status == 'completed' for o in all_orders)

    combined_subtotal = Decimal('0')
    orders_data = []
    for o in all_orders:
        items_data = []
        order_total = Decimal('0')
        for item in o.items.select_related('menu_item').all():
            items_data.append({
                'name': item.menu_item.name,
                'qty':  item.quantity,
                'total': float(item.total_price),
            })
            order_total += item.total_price
            combined_subtotal += item.total_price
        orders_data.append({'items': items_data, 'total': float(order_total)})

    disc_pct = root.discount_percent
    disc_amt = (combined_subtotal * disc_pct / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if disc_pct else Decimal('0')
    combined_total = combined_subtotal - disc_amt

    return JsonResponse({
        'order_id':       root.id,
        'order_number':   root.order_number,
        'status':         root.status,
        'all_completed':  all_completed,
        'customer_name':  root.customer_name,
        'customer_phone': root.customer_phone,
        'combined_total': float(combined_total),
        'orders':         orders_data,
    })


def combo_detail_api(request, combo_id):
    """JSON API — returns combo details with items for customer modal."""
    try:
        combo = Combo.objects.prefetch_related('combo_items__menu_item').get(id=combo_id, is_active=True)
    except Combo.DoesNotExist:
        return JsonResponse({'error': 'not found'}, status=404)
    items_data = []
    for ci in combo.combo_items.all():
        mi = ci.menu_item
        items_data.append({
            'name': mi.name,
            'qty': ci.quantity,
            'price': float(mi.price),
            'image': mi.image.url if mi.image else None,
            'type': mi.item_type,
        })
    return JsonResponse({
        'id': combo.id,
        'name': combo.name,
        'description': combo.description,
        'price': float(combo.price),
        'icon': combo.icon or '🎁',
        'image': combo.image.url if combo.image else None,
        'items': items_data,
    })


@login_required(login_url='/staff/login/')
def staff_portal(request):
    from django.db.models import Count, Avg
    tables = Table.objects.filter(is_active=True).order_by('number')
    pending_orders  = Order.objects.filter(status='pending').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    accepted_orders = Order.objects.filter(status='accepted').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    cooking_orders  = Order.objects.filter(status='preparing').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    ready_orders    = Order.objects.filter(status='ready').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    completed_orders = Order.objects.filter(status='completed').select_related('table').prefetch_related('items__menu_item').order_by('-created_at')[:30]

    today = timezone.now().date()
    week_start  = today - timezone.timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    # Today stats
    today_qs = Order.objects.filter(status='completed', created_at__date=today)
    today_revenue       = today_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    today_orders_count  = Order.objects.filter(created_at__date=today).count()
    today_completed     = today_qs.count()

    # Weekly stats
    week_qs      = Order.objects.filter(status='completed', created_at__date__gte=week_start)
    week_revenue = week_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    week_orders  = week_qs.count()

    # Monthly stats
    month_qs      = Order.objects.filter(status='completed', created_at__date__gte=month_start)
    month_revenue = month_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    month_orders  = month_qs.count()

    # All-time stats
    all_qs        = Order.objects.filter(status='completed')
    all_revenue   = all_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    all_orders    = all_qs.count()
    avg_order_val = all_qs.aggregate(avg=Avg('total_amount'))['avg'] or Decimal('0')

    # Top 8 selling items (by quantity sold this month)
    from django.db.models import F
    top_items = (
        OrderItem.objects
        .filter(order__status='completed', order__created_at__date__gte=month_start)
        .values(item_name=F('menu_item__name'))
        .annotate(total_qty=Sum('quantity'), total_rev=Sum('total_price'))
        .order_by('-total_qty')[:8]
    )

    # Last 7 days revenue for chart
    from datetime import timedelta
    chart_labels = []
    chart_values = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        rev = Order.objects.filter(status='completed', created_at__date=d).aggregate(t=Sum('total_amount'))['t'] or 0
        chart_labels.append(d.strftime('%d %b'))
        chart_values.append(float(rev))

    shop = get_shop()
    discounts = Discount.objects.filter(is_active=True)
    pending_ids = list(pending_orders.values_list('id', flat=True))

    return render(request, 'restaurant/staff_portal.html', {
        'tables': tables,
        'pending_orders': pending_orders,
        'accepted_orders': accepted_orders,
        'cooking_orders': cooking_orders,
        'ready_orders': ready_orders,
        'completed_orders': completed_orders,
        'today_revenue': today_revenue,
        'today_orders_count': today_orders_count,
        'today_completed': today_completed,
        'week_revenue': week_revenue,
        'week_orders': week_orders,
        'month_revenue': month_revenue,
        'month_orders': month_orders,
        'all_revenue': all_revenue,
        'all_orders': all_orders,
        'avg_order_val': avg_order_val,
        'top_items': top_items,
        'chart_labels_json': json.dumps(chart_labels),
        'chart_values_json': json.dumps(chart_values),
        'shop': shop,
        'discounts': discounts,
        'pending_ids_json': json.dumps(pending_ids),
    })


@login_required(login_url='/staff/login/')
@require_POST
def update_order_status(request, order_id):
    try:
        order = get_object_or_404(Order, id=order_id)
        data = json.loads(request.body)
        new_status = data.get('status')
        payment_method = data.get('payment_method', '')
        discount_percent = data.get('discount_percent')

        if discount_percent is not None:
            order.discount_percent = Decimal(str(discount_percent))

        valid_transitions = {
            'pending':   ['accepted', 'cancelled'],
            'accepted':  ['preparing', 'cancelled'],
            'preparing': ['ready'],
            'ready':     ['completed'],
            'completed': [], 'cancelled': [],
        }
        if new_status not in valid_transitions.get(order.status, []):
            return JsonResponse({'success': False, 'error': f'Cannot move from {order.status} to {new_status}'})

        order.status = new_status
        if new_status == 'completed':
            order.completed_at = timezone.now()
            if payment_method == 'offline':
                order.payment_status = 'paid_offline'
                order.payment_method = 'offline'
            elif payment_method == 'online':
                order.payment_status = 'paid_online'
                order.payment_method = 'online'
        order.calculate_totals()
        order.save()
        return JsonResponse({'success': True, 'status': order.status})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def pending_orders_api(request):
    """Polling endpoint — returns new pending orders since last_id"""
    try:
        last_id = int(request.GET.get('last_id', 0))
        orders = Order.objects.filter(id__gt=last_id, status='pending').select_related('table')
        result = []
        for o in orders:
            result.append({
                'id': o.id,
                'order_number': o.order_number,
                'customer_name': o.customer_name,
                'table': o.table.name if o.table else 'Takeaway',
                'total': float(o.total_amount),
                'order_type': o.order_type,
            })
        return JsonResponse({'orders': result})
    except Exception as e:
        return JsonResponse({'orders': [], 'error': str(e)})


# ─── EXCEL EXPORT ─────────────────────────────────────────────────

@login_required(login_url='/staff/login/')
def export_excel(request):
    try:
        orders = Order.objects.all().select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        wb = _build_excel(orders, 'All Orders')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="brothers_cafe_all_orders.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)


@login_required(login_url='/staff/login/')
def export_daily_excel(request):
    try:
        today = timezone.now().date()
        orders = Order.objects.filter(created_at__date=today).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        wb = _build_excel(orders, f'Orders {today}')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_{today}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)






@login_required(login_url='/staff/login/')
def download_qr_codes(request):
    from reportlab.graphics.barcode.qr import QrCodeWidget
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderSVG
    import re, os, tempfile

    base_url = request.build_absolute_uri('/').rstrip('/')

    tables_data = [
        {"label": t.name, "url": f"{base_url}/table/{t.id}/menu/", "takeaway": False}
        for t in Table.objects.filter(is_active=True).order_by('number')
    ]
    tables_data.append({"label": "Takeaway", "url": f"{base_url}/takeaway/", "takeaway": True})

    def make_qr_svg(url, size=210):
        qr = QrCodeWidget(url)
        b = qr.getBounds()
        d = Drawing(b[2]-b[0], b[3]-b[1])
        d.add(qr)
        tmp = tempfile.mktemp(suffix='.svg')
        renderSVG.drawToFile(d, tmp)
        with open(tmp) as f:
            svg = f.read()
        os.remove(tmp)
        svg = re.sub(r'<\?xml[^>]+\?>', '', svg)
        svg = re.sub(r'<!DOCTYPE[^>]+>', '', svg)
        svg = re.sub(r'width="[^"]+"', f'width="{size}"', svg, count=1)
        svg = re.sub(r'height="[^"]+"', f'height="{size}"', svg, count=1)
        return svg.strip()

    cards = []
    for t in tables_data:
        svg    = make_qr_svg(t["url"])
        color  = "#e67e22" if t["takeaway"] else "#1a1a2e"
        accent = "#fff8f0" if t["takeaway"] else "#f4f6ff"
        icon   = "🛍️" if t["takeaway"] else "🍽️"
        sub    = "Scan to order takeaway" if t["takeaway"] else "Scan to order food"
        cards.append(f"""<div class="qr-card">
<div class="card-top" style="background:{color}">
  <div class="card-icon">{icon}</div>
  <div><div class="card-name">Brothers Cafe</div><div class="card-sub">{sub}</div></div>
</div>
<div class="card-qr" style="background:{accent}">{svg}</div>
<div class="card-bottom" style="background:{color}">
  <span class="scan-text">☕ Scan &amp; Order</span>
  <span class="staff-label">Staff: {t['label']}</span>
</div></div>""")

    html = '''<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Brothers Cafe QR Codes</title>
<style>
@page{margin:15mm}
@media print{.no-print{display:none!important}body{background:white;padding:0}}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#eef0f3;padding:28px 20px}
h1{text-align:center;font-size:24px;font-weight:800;color:#1a1a2e;margin-bottom:6px}
.sub{text-align:center;font-size:13px;color:#666;margin-bottom:26px}
.grid{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;max-width:820px;margin:0 auto}
.qr-card{background:white;border-radius:16px;overflow:hidden;box-shadow:0 4px 18px rgba(0,0,0,.13);break-inside:avoid}
.card-top{padding:14px 16px;display:flex;align-items:center;gap:10px;color:white}
.card-icon{font-size:26px}.card-name{font-size:14px;font-weight:800}.card-sub{font-size:11px;opacity:.75;margin-top:2px}
.card-qr{padding:16px;display:flex;align-items:center;justify-content:center}
.card-bottom{padding:10px 16px;color:white;display:flex;align-items:center;justify-content:space-between}
.scan-text{font-size:13px;font-weight:700}
.staff-label{font-size:10px;background:rgba(255,255,255,.18);padding:3px 9px;border-radius:20px}
.print-btn{display:block;margin:24px auto 0;padding:13px 38px;background:linear-gradient(135deg,#1a1a2e,#0f3460);color:white;border:none;border-radius:50px;font-size:15px;font-weight:700;cursor:pointer;font-family:inherit}
</style></head><body>
<h1>☕ Brothers Cafe — QR Codes</h1>
<p class="sub">Print &amp; place on each table. Customers scan to order — table numbers never shown to customers.</p>
<div class="grid">''' + ''.join(cards) + '''</div>
<button class="print-btn no-print" onclick="window.print()">🖨️ Print All QR Codes</button>
</body></html>'''

    return HttpResponse(html, content_type='text/html; charset=utf-8')

@login_required(login_url='/staff/login/')
def export_weekly_excel(request):
    try:
        today = timezone.now().date()
        week_start = today - timezone.timedelta(days=today.weekday())
        orders = Order.objects.filter(created_at__date__gte=week_start).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        wb = _build_excel(orders, f'Weekly {week_start} to {today}')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_weekly_{week_start}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)


@login_required(login_url='/staff/login/')
def export_monthly_excel(request):
    try:
        today = timezone.now().date()
        month_start = today.replace(day=1)
        orders = Order.objects.filter(created_at__date__gte=month_start).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        wb = _build_excel(orders, f'Monthly {today.strftime("%B %Y")}')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_{today.strftime("%Y_%m")}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)


@login_required(login_url='/staff/login/')
def export_range_excel(request):
    try:
        from datetime import datetime
        date_from_str = request.GET.get('from', '')
        date_to_str = request.GET.get('to', '')
        if not date_from_str or not date_to_str:
            return HttpResponse('Missing from/to parameters', status=400)
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        orders = Order.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        title = f'{date_from.strftime("%d %b")} to {date_to.strftime("%d %b %Y")}'
        wb = _build_excel(orders, title)
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_{date_from_str}_to_{date_to_str}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)

def _build_excel(orders, title):
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = f'Brothers Cafe — {title}'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='1a1a2e')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Order #','Type','Date','Time','Customer','Phone','Table',
               'Items','Subtotal (Rs)','Disc%','Disc (Rs)','Total (Rs)','Status','Payment']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='e74c3c')
        c.alignment = Alignment(horizontal='center')

    row = 3
    total_rev = Decimal('0')
    for order in orders:
        items_str = ', '.join(f"{i.quantity}x {i.menu_item.name}" for i in order.items.all())
        ldt = timezone.localtime(order.created_at)
        ws.cell(row=row, column=1,  value=order.order_number)
        ws.cell(row=row, column=2,  value=order.get_order_type_display())
        ws.cell(row=row, column=3,  value=ldt.strftime('%d-%m-%Y'))
        ws.cell(row=row, column=4,  value=ldt.strftime('%H:%M'))
        ws.cell(row=row, column=5,  value=order.customer_name)
        ws.cell(row=row, column=6,  value=order.customer_phone)
        ws.cell(row=row, column=7,  value=str(order.table) if order.table else 'Takeaway')
        ws.cell(row=row, column=8,  value=items_str)
        ws.cell(row=row, column=9,  value=float(order.subtotal))
        ws.cell(row=row, column=10, value=float(order.discount_percent))
        ws.cell(row=row, column=11, value=float(order.discount_amount))
        ws.cell(row=row, column=12, value=float(order.total_amount))
        ws.cell(row=row, column=13, value=order.get_status_display())
        ws.cell(row=row, column=14, value=order.get_payment_status_display())
        if order.status == 'completed':
            total_rev += order.total_amount
        row += 1

    row += 1
    ws.cell(row=row, column=11, value='Total Revenue:').font = Font(bold=True)
    ws.cell(row=row, column=12, value=float(total_rev)).font = Font(bold=True)

    # Set column widths using letters directly (avoids MergedCell.column_letter error)
    col_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']
    for letter in col_letters:
        ws.column_dimensions[letter].width = 16
    ws.column_dimensions['H'].width = 45  # Items column wider
    return wb


@login_required(login_url='/staff/login/')
def pos_terminal(request):
    """Staff-facing POS Terminal — browse categories/items, build & save orders."""
    categories = Category.objects.filter(is_active=True).prefetch_related(
        Prefetch('items', queryset=MenuItem.objects.filter(is_available=True).order_by('order', 'name'))
    )
    discounts = Discount.objects.filter(is_active=True)
    shop = get_shop()
    return render(request, 'restaurant/pos_terminal.html', {
        'categories': categories,
        'discounts': discounts,
        'shop': shop,
    })
